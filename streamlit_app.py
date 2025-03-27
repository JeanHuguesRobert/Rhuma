# streamlit_app.py
#   Rhuma, rhum solaire de Corse
#
# (c) Jean Hugues Robert, 03/2025
# MIT License

import streamlit as st
import pandas as pd
import matplotlib.pyplot as plt
import os
import json
from datetime import datetime
import gspread
from oauth2client.service_account import ServiceAccountCredentials
import numpy as np
from dotenv import load_dotenv
import zipfile
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import tempfile
import shutil
from modules.data_export import export_to_google_sheets, export_to_excel, export_to_json
from modules.pvgis_analysis import pvgis_analysis_section
from modules.rhuma_state import rhuma, state_manager

# Charger les variables d'environnement
load_dotenv()

# Initialiser l'état RHUMA avec les valeurs par défaut
RHUMA = state_manager.get_state()

# Paramètres de base
RHUMA_ID = RHUMA["metadata"]["id"]
RHUMA_LABEL = RHUMA["metadata"]["label"]

# Constantes de configuration
SURFACE_CANNE_MIN = 1000  # m²
SURFACE_CANNE_MAX = 5000  # m²
SURFACE_CANNE_DEFAULT = RHUMA["configuration"]["surface_canne"]  # m²

RENDEMENT_CANNE_MIN = 80  # t/ha
RENDEMENT_CANNE_MAX = 160  # t/ha
RENDEMENT_CANNE_DEFAULT = RHUMA["configuration"]["rendement_canne"]  # t/ha

TENEUR_SUCRE_MIN = 10  # %
TENEUR_SUCRE_MAX = 20  # %
TENEUR_SUCRE_DEFAULT = RHUMA["configuration"]["teneur_sucre"]  # %

EFFICACITE_EXTRACTION_MIN = 60  # %
EFFICACITE_EXTRACTION_MAX = 95  # %
EFFICACITE_EXTRACTION_DEFAULT = RHUMA["configuration"]["efficacite_extraction"]  # %

EFFICACITE_DISTILLATION_MIN = 60  # %
EFFICACITE_DISTILLATION_MAX = 95  # %
EFFICACITE_DISTILLATION_DEFAULT = RHUMA["configuration"]["efficacite_distillation"]  # %

PV_SERRE_MAX = 500  # kWc
PV_SERRE_DEFAULT = RHUMA["configuration"]["pv_serre"]  # kWc

PV_SOL_MAX = 500  # kWc
PV_SOL_DEFAULT = RHUMA["configuration"]["pv_sol"]  # kWc

TARIF_S24_MIN = 0.05  # €/kWh
TARIF_S24_MAX = 0.20  # €/kWh
TARIF_S24_DEFAULT = RHUMA["configuration"]["tarif_s24"]  # €/kWh

TVA_MIN = 0  # %
TVA_MAX = 20  # %
TVA_DEFAULT = RHUMA["configuration"]["tva"]  # %

LIMITE_PUISSANCE_S24 = 500  # kWc
HEURES_PLEIN_SOLEIL = 1600  # heures/an
LIMITE_PRODUCTION_S24 = LIMITE_PUISSANCE_S24 * HEURES_PLEIN_SOLEIL  # kWh/an
TARIF_S24_DEPASSEMENT = 0.05  # €/kWh au-delà de la limite

class TrackingSystemSimulation:
    def __init__(self, standard_panel_efficiency=0.22):
        self.base_efficiency = standard_panel_efficiency
        
    def calculate_tracking_gains(self, tracking_precision=0.2, panel_orientation_precision=2):
        """
        Calcule les gains de production dus au tracking
        
        Args:
        - tracking_precision (float): Précision du tracking en degrés
        - panel_orientation_precision (float): Précision de l'orientation du panneau
        
        Returns:
        - Dict avec gains et détails de performance
        """
        # Nouveaux gains de production basés sur les données réelles
        base_tracking_gain = {
            "single_axis": 0.279,  # 27.9% de gain avec tracking monoaxe (par rapport à fixe)
            "dual_axis": 0.279,    # 27.9% de gain avec tracking monoaxe (par rapport à fixe)
        }
        
        # Impact de la précision sur le gain
        precision_impact = {
            "low": -0.05,   # précision > 1° : perte de 5%
            "medium": -0.02, # précision 0.5-1° : perte de 2%
            "high": 0.00    # précision < 0.5° : pas de perte
        }
        
        precision_category = (
            "high" if tracking_precision < 0.5 else
            "medium" if tracking_precision < 1 else
            "low"
        )
        
        gains = {
            "single_axis_gain": max(0, base_tracking_gain["single_axis"] + precision_impact[precision_category]),
            "dual_axis_gain": max(0, base_tracking_gain["dual_axis"] + precision_impact[precision_category])
        }
        
        # Impact de l'orientation du panneau
        orientation_loss = panel_orientation_precision * 0.01  # 1% de perte par degré d'orientation
        
        for key in gains:
            gains[key] = max(0, gains[key] - orientation_loss)
        
        # Ajout des pourcentages de référence
        gains["fixed_reference"] = 1.0
        gains["tracking_reference"] = 1.279  # 127.9% de la production fixe
        gains["tracking_loss"] = 0.218  # 21.8% de perte par rapport au tracking idéal
        
        return gains
    
    def simulate_microinverter_performance(self, input_voltage=40, output_voltage=230):
        """
        Simulation des performances du micro-onduleur
        
        Args:
        - input_voltage (float): Tension d'entrée DC
        - output_voltage (float): Tension de sortie AC
        
        Returns:
        - Dict avec caractéristiques de performance
        """
        conversion_efficiency = min(0.96, 1 - abs(input_voltage - 40)/100)
        
        return {
            "max_power": 350,  # Watts
            "conversion_efficiency": conversion_efficiency,
            "input_voltage_range": [30, 60],
            "output_voltage": output_voltage,
            "temperature_coefficient": -0.4  # % par °C
        }
    
    def monte_carlo_solar_production(self, base_production, num_simulations=1000):
        """
        Simulation Monte Carlo des variations de production
        
        Args:
        - base_production (float): Production de base
        - num_simulations (int): Nombre de simulations
        
        Returns:
        - Dict avec statistiques de production
        """
        np.random.seed(42)
        
        # Variations possibles
        variations = np.random.normal(1, 0.05, num_simulations)
        productions = base_production * variations
        
        return {
            "mean_production": np.mean(productions),
            "median_production": np.median(productions),
            "std_deviation": np.std(productions),
            "min_production": np.min(productions),
            "max_production": np.max(productions)
        }


def get_pv_production_data():
    """
    Returns PV production data based on PVGIS calculations for 500kW system
    """
    # Monthly production data from PVGIS (in kWh)
    monthly_pv_production = {
        1: 56182.35,  # January
        2: 60184.96,  # February
        3: 74890.22,  # March
        4: 80076.42,  # April
        5: 94987.89,  # May
        6: 100008.61, # June
        7: 108980.42, # July
        8: 97314.88,  # August
        9: 78784.55,  # September
        10: 69051.25, # October
        11: 50725.67, # November
        12: 55914.12  # December
    }
    
    # Total annual production from PVGIS
    total_pv_production = 927101.33  # kWh
    
    return monthly_pv_production, total_pv_production

def tracking_optimization_section(production_pv_ideal):
    st.header("🔍 Optimisation du Tracking Solaire")
    
    # Get PVGIS data
    monthly_pv_production, total_pv_production = get_pv_production_data()
    
    # Initialisation du simulateur de tracking
    tracker = TrackingSystemSimulation()
    
    # Colonnes pour les paramètres
    col1, col2, col3 = st.columns(3)
    
    with col1:
        tracking_strategy = st.selectbox(
            "Stratégie de Tracking", 
            ["Statique", "Single-Axis", "Dual-Axis"],
            help="Sélectionnez la stratégie de suivi solaire"
        )
    
    with col2:
        tracking_precision = st.slider(
            "Précision de Tracking", 
            0.1, 5.0, 0.2,
            help="Précision angulaire du système de tracking (en degrés)"
        )
    
    with col3:
        panel_orientation = st.slider(
            "Précision d'Orientation", 
            0.5, 5.0, 2.0,
            help="Précision de l'orientation initiale du panneau"
        )
    
    # Calcul des gains de tracking
    tracking_gains = tracker.calculate_tracking_gains(
        tracking_precision, 
        panel_orientation
    )
    
    # Simulation des performances du micro-onduleur
    microinverter_perf = tracker.simulate_microinverter_performance()
    
    # Création de colonnes pour afficher les résultats
    st.subheader("📊 Résultats de l'Optimisation")
    
    gains_col1, gains_col2, gains_col3 = st.columns(3)
    
    with gains_col1:
        st.metric(
            "Production Fixe (Référence)", 
            f"{total_pv_production/1000:.2f} MWh",
            help="Production de référence sans tracking"
        )
    
    with gains_col2:
        st.metric(
            "Production avec Tracking", 
            f"{total_pv_production/1000 * (1 + tracking_gains['single_axis_gain']):.2f} MWh",
            help="Production avec tracking monoaxe"
        )
    
    with gains_col3:
        st.metric(
            "Gain Tracking", 
            f"{tracking_gains['single_axis_gain']*100:.1f}%",
            help="Augmentation de production par rapport à la production fixe"
        )
    
    # Section Analyse de Production
    st.subheader("📈 Analyse de Production")
    
    # Création du graphique des productions mensuelles
    months = list(monthly_pv_production.keys())
    monthly_values = list(monthly_pv_production.values())
    
    fig, ax = plt.subplots()
    ax.bar(months, monthly_values, color="#4CAF50")
    ax.set_xticks(months)
    ax.set_xticklabels(["Jan", "Fév", "Mar", "Avr", "Mai", "Juin", "Juil", "Août", "Sep", "Oct", "Nov", "Déc"])
    ax.set_xlabel("Mois")
    ax.set_ylabel("Production (kWh)")
    ax.set_title("Production PV mensuelle")
    
    st.pyplot(fig)
    
    # Section des pertes
    st.subheader("📊 Détails des Pertes")
    
    # Données des pertes de PVGIS
    losses = {
        "Angle d'incidence": -1.44,
        "Spectre": 1.23,
        "Température et irradiance": -9.54,
        "Total": -18.77
    }
    
    # Création du graphique des pertes
    fig2, ax2 = plt.subplots()
    ax2.bar(losses.keys(), losses.values(), color="#FFC107")
    ax2.axhline(0, color="black", linewidth=0.5)
    ax2.set_ylabel("Pourcentage de perte (%)")
    ax2.set_title("Répartition des pertes")
    
    st.pyplot(fig2)
    
    # Paramètres avancés
    with st.expander("🛠️ Paramètres Techniques Avancés"):
        st.write("### Caractéristiques du Système PV")
        st.write(f"- Puissance installée : 500 kW")
        st.write(f"- Technologie : c-Si")
        st.write(f"- Perte système : 10%")
        st.write(f"- Source de données : PVGIS (2005-2023)")
        st.write(f"- Base de données : SARAH3")
        st.write(f"- Base météo : ERA5")
        st.write(f"- Altitude : 497 m")
        
        st.write("\n### Performance Annuelle")
        st.write(f"- Production annuelle : {total_pv_production/1000:.2f} MWh")
        st.write(f"- Irradiation annuelle : 2282.62 kWh/m²")
        st.write(f"- Écart-type annuel : {39070.05/1000:.2f} MWh")


def tracking_comparison_section(production_pv_ideal):
    st.header("📊 Comparaison Système Fixe vs Double Axe")
    
    # Get PVGIS data
    monthly_pv_production, total_pv_production = get_pv_production_data()
    
    # Initialisation du simulateur de tracking
    tracker = TrackingSystemSimulation()
    
    # Calcul des gains de tracking pour différentes précisions
    tracking_precisions = [0.2, 1.0, 2.0]  # Précisions en degrés
    results = []
    
    for precision in tracking_precisions:
        # Calcul des gains avec cette précision
        tracking_gains = tracker.calculate_tracking_gains(precision, precision)
        
        # Production avec tracking
        production_tracking = total_pv_production * (1 + tracking_gains['single_axis_gain'])
        
        # Calcul des CA
        ca_fixe = calcul_tarif_production(total_pv_production)
        ca_tracking = calcul_tarif_production(production_tracking)
        
        # Calcul des gains en CA
        gain_ca = ((ca_tracking - ca_fixe) / ca_fixe) * 100
        
        results.append({
            'precision': precision,
            'production_tracking': production_tracking,
            'ca_fixe': ca_fixe,
            'ca_tracking': ca_tracking,
            'gain_ca': gain_ca
        })
    
    # Création du tableau de comparaison
    st.subheader("📈 Comparaison des Performances")
    
    # Conversion des résultats en DataFrame pour affichage
    df_results = pd.DataFrame(results)
    df_results['production_tracking'] = df_results['production_tracking'].apply(lambda x: f"{x/1000:.2f} MWh")
    df_results['ca_fixe'] = df_results['ca_fixe'].apply(lambda x: f"{x/1000:.2f} k€")
    df_results['ca_tracking'] = df_results['ca_tracking'].apply(lambda x: f"{x/1000:.2f} k€")
    df_results['gain_ca'] = df_results['gain_ca'].apply(lambda x: f"{x:.1f}%")
    
    # Renommer les colonnes pour une meilleure lisibilité
    df_results = df_results.rename(columns={
        'precision': 'Précision (°)',
        'production_tracking': 'Production avec Tracking (MWh)',
        'ca_fixe': 'CA Système Fixe (k€)',
        'ca_tracking': 'CA avec Tracking (k€)',
        'gain_ca': 'Gain en CA (%)'
    })
    
    st.dataframe(df_results, hide_index=True)
    
    # Création des graphiques comparatifs
    st.subheader("📊 Analyse Graphique")
    
    # Graphique de la production
    fig1, ax1 = plt.subplots(figsize=(10, 6))
    
    # Production fixe
    ax1.bar(['Système Fixe'], [total_pv_production/1000], color='#4CAF50', label='Système Fixe')
    
    # Productions avec tracking pour différentes précisions
    for i, precision in enumerate(tracking_precisions):
        production = df_results.loc[df_results['Précision (°)'] == precision, 'Production avec Tracking (MWh)'].values[0]
        ax1.bar([f'Tracking {precision}°'], [float(production.split()[0])], 
                color=f'C{i+1}', label=f'Tracking {precision}°')
    
    ax1.set_ylabel('Production annuelle (MWh)')
    ax1.set_title('Comparaison de la Production Annuelle')
    ax1.legend()
    st.pyplot(fig1)
    
    # Graphique du CA
    fig2, ax2 = plt.subplots(figsize=(10, 6))
    
    # CA fixe
    ca_fixe = float(df_results['CA Système Fixe (k€)'].iloc[0].split()[0])
    ax2.bar(['Système Fixe'], [ca_fixe], color='#4CAF50', label='Système Fixe')
    
    # CA avec tracking pour différentes précisions
    for i, precision in enumerate(tracking_precisions):
        ca = df_results.loc[df_results['Précision (°)'] == precision, 'CA avec Tracking (k€)'].values[0]
        ax2.bar([f'Tracking {precision}°'], [float(ca.split()[0])], 
                color=f'C{i+1}', label=f'Tracking {precision}°')
    
    ax2.set_ylabel('Chiffre d\'Affaires annuel (k€)')
    ax2.set_title('Comparaison du Chiffre d\'Affaires Annuel')
    ax2.legend()
    st.pyplot(fig2)
    
    # Section de conclusion
    with st.expander("📝 Conclusion"):
        st.write("### Analyse des Résultats")
        st.write("""
        - Le système à double axe tracking permet une augmentation significative de la production
        - Les gains en CA sont proportionnels aux gains en production
        - La précision du système de tracking a un impact direct sur les performances
        - Un système de tracking précis (0.2°) peut augmenter le CA de plus de 25%
        """)
        
        st.write("### Recommandations")
        st.write("""
        - Investir dans un système de tracking précis si l'espace est limité
        - Considérer le coût supplémentaire du tracking par rapport aux gains
        - Prendre en compte la maintenance du système de tracking
        """)

def calcul_tarif_production(production_kwh):
    """
    Calcule le revenu en tenant compte du seuil des 1600h à pleine puissance
    """
    if production_kwh <= LIMITE_PRODUCTION_S24:
        return production_kwh * 0.1772
    else:
        return (LIMITE_PRODUCTION_S24 * 0.1772 + 
                (production_kwh - LIMITE_PRODUCTION_S24) * 0.05)

# Configuration de la page
st.set_page_config(page_title="Simulateur Rhuma, rhum solaire en Corse", layout="wide")
st.title(f"{RHUMA_LABEL}")

# Lecture et affichage des documents Markdown
DOCS_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "docs")

def read_markdown_file(markdown_file):
    """Lire et retourner le contenu d'un fichier Markdown."""
    try:
        with open(markdown_file, 'r', encoding='utf-8') as f:
            return f.read()
    except Exception as e:
        return f"Erreur lors de la lecture du fichier: {str(e)}"

st.markdown("""
## 📚 Documentation

- [Guide Utilisateur](docs/user_guide.md)
- [Architecture Technique](docs/ARCHITECTURE.md)
- [Guide de Déploiement](docs/DEPLOYMENT.md)
- [Roadmap](docs/ROADMAP.md) - Vision stratégique du projet et objectifs à long terme
- [TODO](docs/TODO.md) - Liste des tâches opérationnelles et en cours

[GitHub Repository](https://github.com/JeanHuguesRobert/Rhuma)
""")

# Créer des onglets pour la documentation
doc_tabs = st.tabs(["Crowdfunding", "Documentation technique", "Guide utilisateur"])

with doc_tabs[0]:
    crowdfunding_content = read_markdown_file(os.path.join(DOCS_DIR, "crowdfunding.md"))
    st.markdown(crowdfunding_content)

with doc_tabs[1]:
    technical_content = read_markdown_file(os.path.join(DOCS_DIR, "ARCHITECTURE.md"))
    st.markdown(technical_content)

with doc_tabs[2]:
    guide_content = read_markdown_file(os.path.join(DOCS_DIR, "user_guide.md"))
    st.markdown(guide_content)

def simulate_financial_scenarios(production_fixe, production_tracking):
    """
    Simule les trois scénarios financiers et retourne les résultats
    """
    # Scénario 1: Revente à EDF au tarif S24
    scenario_1 = {
        'nom': 'Revente EDF S24',
        'fixe': {
            'production': production_fixe,
            'autoconsommation': autoconsommation_fixe,
            'revente': production_fixe - autoconsommation_fixe,
            'revenu': (production_fixe - autoconsommation_fixe) * tarif_s24,
            'cout_total': cout_fixe + cout_maintenance + cout_assurance + cout_production
        },
        'tracking': {
            'production': production_tracking,
            'autoconsommation': autoconsommation_tracking,
            'revente': production_tracking - autoconsommation_tracking,
            'revenu': (production_tracking - autoconsommation_tracking) * tarif_s24,
            'cout_total': cout_fixe + cout_tracking + cout_maintenance + cout_assurance + cout_production
        }
    }

    # Scénario 2: Autoconsommation collective
    scenario_2 = {
        'nom': 'Autoconsommation Collective',
        'fixe': {
            'production': production_fixe,
            'autoconsommation': production_fixe,
            'revente': 0,
            'revenu': production_fixe * tarif_heures_creuses,
            'cout_total': cout_fixe + cout_maintenance + cout_assurance + cout_production
        },
        'tracking': {
            'production': production_tracking,
            'autoconsommation': production_tracking,
            'revente': 0,
            'revenu': production_tracking * tarif_heures_creuses,
            'cout_total': cout_fixe + cout_tracking + cout_maintenance + cout_assurance + cout_production
        }
    }

    # Scénario 3: Mixte (autoconsommation + revente)
    scenario_3 = {
        'nom': 'Mixte (Autoconsommation + Revente)',
        'fixe': {
            'production': production_fixe,
            'autoconsommation': autoconsommation_fixe,
            'revente': production_fixe - autoconsommation_fixe,
            'revenu': (autoconsommation_fixe * tarif_heures_creuses + 
                      (production_fixe - autoconsommation_fixe) * tarif_s24),
            'cout_total': cout_fixe + cout_maintenance + cout_assurance + cout_production
        },
        'tracking': {
            'production': production_tracking,
            'autoconsommation': autoconsommation_tracking,
            'revente': production_tracking - autoconsommation_tracking,
            'revenu': (autoconsommation_tracking * tarif_heures_creuses + 
                      (production_tracking - autoconsommation_tracking) * tarif_s24),
            'cout_total': cout_fixe + cout_tracking + cout_maintenance + cout_assurance + cout_production
        }
    }

    # Calcul du ROI pour chaque scénario
    scenarios = [scenario_1, scenario_2, scenario_3]
    for scenario in scenarios:
        for system in ['fixe', 'tracking']:
            # Calcul du bénéfice annuel
            benefice_annuel = scenario[system]['revenu'] - scenario[system]['cout_total']
            
            # Calcul du ROI
            investissement = scenario[system]['cout_total']
            if benefice_annuel > 0:
                roi = (benefice_annuel / investissement) * 100
                temps_retour = investissement / benefice_annuel
            else:
                roi = 0
                temps_retour = float('inf')
            
            scenario[system]['benefice_annuel'] = benefice_annuel
            scenario[system]['roi'] = roi
            scenario[system]['temps_retour'] = temps_retour

    return scenarios

def financial_simulation_section():
    st.header("📊 Simulation Financière")
    
    # Get PVGIS data
    monthly_pv_production, total_pv_production = get_pv_production_data()
    
    # Initialisation du simulateur de tracking
    tracker = TrackingSystemSimulation()
    
    # Calcul des gains de tracking
    tracking_gains = tracker.calculate_tracking_gains(precision_tracking, precision_tracking)
    production_tracking = total_pv_production * (1 + tracking_gains['single_axis_gain'])
    
    # Simulation des scénarios
    scenarios = simulate_financial_scenarios(total_pv_production, production_tracking)
    
    # Création des tableaux de comparaison
    st.subheader("📈 Comparaison des Scénarios")
    
    # Tableau comparatif
    comparison_data = []
    for scenario in scenarios:
        comparison_data.append({
            'Scénario': scenario['nom'],
            'Système': 'Fixe',
            'Production (MWh)': f"{total_pv_production/1000:.2f}",
            'Autoconsommation (MWh)': f"{scenario['fixe']['autoconsommation']/1000:.2f}",
            'Revente (MWh)': f"{scenario['fixe']['revente']/1000:.2f}",
            'Revenu annuel (k€)': f"{scenario['fixe']['revenu']/1000:.2f}",
            'Bénéfice annuel (k€)': f"{scenario['fixe']['benefice_annuel']/1000:.2f}",
            'ROI (%)': f"{scenario['fixe']['roi']:.1f}",
            'Temps retour (ans)': f"{scenario['fixe']['temps_retour']:.1f}"
        })
        comparison_data.append({
            'Scénario': scenario['nom'],
            'Système': 'Tracking',
            'Production (MWh)': f"{production_tracking/1000:.2f}",
            'Autoconsommation (MWh)': f"{scenario['tracking']['autoconsommation']/1000:.2f}",
            'Revente (MWh)': f"{scenario['tracking']['revente']/1000:.2f}",
            'Revenu annuel (k€)': f"{scenario['tracking']['revenu']/1000:.2f}",
            'Bénéfice annuel (k€)': f"{scenario['tracking']['benefice_annuel']/1000:.2f}",
            'ROI (%)': f"{scenario['tracking']['roi']:.1f}",
            'Temps retour (ans)': f"{scenario['tracking']['temps_retour']:.1f}"
        })
    
    df_comparison = pd.DataFrame(comparison_data)
    st.dataframe(df_comparison, hide_index=True)
    
    # Création des graphiques comparatifs
    st.subheader("📊 Analyse Graphique")
    
    # Graphique des revenus
    fig1, ax1 = plt.subplots(figsize=(12, 6))
    
    # Revenus par scénario
    for scenario in scenarios:
        ax1.bar([f"{scenario['nom']} - Fixe"], 
                [scenario['fixe']['revenu']/1000], 
                color='#4CAF50',
                alpha=0.6)
        ax1.bar([f"{scenario['nom']} - Tracking"], 
                [scenario['tracking']['revenu']/1000], 
                color='#FFC107',
                alpha=0.6)
    
    ax1.set_ylabel('Revenu annuel (k€)')
    ax1.set_title('Comparaison des Revenus Annuels par Scénario')
    st.pyplot(fig1)
    
    # Graphique des bénéfices
    fig2, ax2 = plt.subplots(figsize=(12, 6))
    
    # Bénéfices par scénario
    for scenario in scenarios:
        ax2.bar([f"{scenario['nom']} - Fixe"], 
                [scenario['fixe']['benefice_annuel']/1000], 
                color='#4CAF50',
                alpha=0.6)
        ax2.bar([f"{scenario['nom']} - Tracking"], 
                [scenario['tracking']['benefice_annuel']/1000], 
                color='#FFC107',
                alpha=0.6)
    
    ax2.set_ylabel('Bénéfice annuel (k€)')
    ax2.set_title('Comparaison des Bénéfices Annuels par Scénario')
    st.pyplot(fig2)
    
    # Graphique des temps de retour
    fig3, ax3 = plt.subplots(figsize=(12, 6))
    
    # Temps de retour par scénario
    for scenario in scenarios:
        ax3.bar([f"{scenario['nom']} - Fixe"], 
                [scenario['fixe']['temps_retour']], 
                color='#4CAF50',
                alpha=0.6)
        ax3.bar([f"{scenario['nom']} - Tracking"], 
                [scenario['tracking']['temps_retour']], 
                color='#FFC107',
                alpha=0.6)
    
    ax3.set_ylabel('Temps de retour (ans)')
    ax3.set_title('Comparaison des Temps de Retour sur Investissement')
    st.pyplot(fig3)
    
    # Section de conclusion
    with st.expander("📝 Analyse des Résultats"):
        st.write("### Synthèse des Scénarios")
        
        # Meilleur scénario par critère
        best_scenarios = {
            'Revenu': max(scenarios, key=lambda x: x['tracking']['revenu']),
            'Bénéfice': max(scenarios, key=lambda x: x['tracking']['benefice_annuel']),
            'ROI': max(scenarios, key=lambda x: x['tracking']['roi']),
            'Temps retour': min(scenarios, key=lambda x: x['tracking']['temps_retour'])
        }
        
        for critere, scenario in best_scenarios.items():
            st.write(f"- Meilleur {critere} : {scenario['nom']} - Tracking")
            st.write(f"  * Revenu : {scenario['tracking']['revenu']/1000:.2f} k€")
            st.write(f"  * Bénéfice : {scenario['tracking']['benefice_annuel']/1000:.2f} k€")
            st.write(f"  * ROI : {scenario['tracking']['roi']:.1f}%")
            st.write(f"  * Temps retour : {scenario['tracking']['temps_retour']:.1f} ans")
            st.write("---")
        
        st.write("### Recommandations")
        st.write("""
        - Le système tracking est toujours plus rentable que le système fixe
        - Le scénario mixte (autoconsommation + revente) offre généralement le meilleur retour sur investissement
        - La durée d'amortissement doit être adaptée aux besoins de financement
        - Les coûts d'exploitation (maintenance, assurance) doivent être soigneusement budgétisés
        """)

def export_to_google_sheets(data, sheet_name="Simulation Rhuma"):
    """
    Exporte les données vers une nouvelle feuille Google Sheets avec une structure optimisée
    """
    try:
        # Initialiser le client
        client = get_google_sheet_client()
        if not client:
            return "Configuration Google Sheets non valide"
            
        # Créer un nouveau spreadsheet
        spreadsheet = client.create(sheet_name)
        
        # Partager le spreadsheet avec l'utilisateur
        spreadsheet.share(
            os.getenv('RHUMA_GOOGLE_SHEETS_CLIENT_EMAIL'),  # Utiliser l'email du service account
            perm_type='user',
            role='writer'
        )
        
        # Créer les différentes feuilles
        config_sheet = spreadsheet.add_worksheet(title="Configuration", rows="200", cols="20")
        results_sheet = spreadsheet.add_worksheet(title="Résultats", rows="200", cols="20")
        calculations_sheet = spreadsheet.add_worksheet(title="Calculs", rows="200", cols="20")
        simulation_sheet = spreadsheet.add_worksheet(title="Simulation", rows="200", cols="20")
        
        # 1. Configuration générale
        config_general = {
            "ID du Projet": RHUMA_ID,
            "Label du Projet": RHUMA_LABEL,
            "Version": os.getenv('RHUMA_VERSION', "1.0.0"),
            "Timestamp": datetime.now().isoformat()
        }
        
        # 2. Paramètres de Simulation
        config_simulation = {
            "Surface canne": rhuma('surface_canne'),
            "Rendement canne": rhuma('rendement_canne'),
            "Teneur sucre": rhuma('teneur_sucre'),
            "Efficacité extraction": rhuma('efficacite_extraction'),
            "Efficacité distillation": rhuma('efficacite_distillation'),
            "Puissance PV (serre)": rhuma('pv_serre'),
            "Puissance PV (au sol)": rhuma('pv_sol'),
            "Tarif S24": rhuma('tarif_s24'),
            "TVA": rhuma('tva'),
            "Coût système PV fixe": rhuma('cout_fixe'),
            "Coût système tracking": rhuma('cout_tracking'),
            "Coût construction serre": rhuma('cout_construction'),
            "Coût maintenance": rhuma('cout_maintenance'),
            "Coût assurance": rhuma('cout_assurance'),
            "Coût production": rhuma('cout_production'),
            "Tarif heures creuses": rhuma('tarif_heures_creuses'),
            "Autoconsommation fixe": rhuma('autoconsommation_fixe'),
            "Autoconsommation tracking": rhuma('autoconsommation_tracking'),
            "Prix du rhum": rhuma('prix_rhum'),
            "Pertes PV": rhuma('pertes_pv'),
            "Pertes tracking": rhuma('pertes_tracking'),
            "Précision tracking": rhuma('precision_tracking'),
            "Taux d'intérêt": rhuma('taux_interet'),
            "Durée d'amortissement": rhuma('duree_amortissement')
        }
        
        # 3. Résultats de la simulation
        results = data
        
        # 4. Calculs intermédiaires
        calculations = {
            "Production PV (serre)": f"=Configuration!B8",
            "Production PV (au sol)": f"=Configuration!B9",
            "Production totale": "=Calculs!B1 + Calculs!B2",
            "Autoconsommation totale": "=Configuration!B10 + Configuration!B11",
            "Revente totale": "=Calculs!B3 - Calculs!B4",
            "Revenu PV": "=Calculs!B5 * Configuration!B12",
            "Revenu Rhum": "=Calculs!B6 * Configuration!B13",
            "Bénéfice net": "=Calculs!B7 - Configuration!B14 - Configuration!B15 - Configuration!B16"
        }
        
        # Créer les données pour chaque feuille
        config_data = [
            ["Configuration Générale"] + [""] * (len(config_general) - 1),
            *[[k, v] for k, v in config_general.items()],
            ["\nParamètres de Simulation"] + [""] * (len(config_simulation) - 1),
            *[[k, v] for k, v in config_simulation.items()]
        ]
        
        results_data = [
            ["Résultats de la Simulation"] + [""] * (len(results) - 1),
            *[[k, v] for k, v in results.items()]
        ]
        
        calculations_data = [
            ["Calculs Intermédiaires"] + [""] * (len(calculations) - 1),
            *[[k, v] for k, v in calculations.items()]
        ]
        
        # Ajouter les formules pour la simulation
        simulation_formulas = [
            ["Paramètre", "Formule", "Description"],
            ["Surface canne", "=Configuration!B2", "Surface totale dédiée à la canne"],
            ["Rendement canne", "=Configuration!B3", "Rendement annuel de la canne"],
            ["Production PV", "=Calculs!B3", "Puissance PV installée"],
            ["Revenu total", "=Calculs!B7", "Revenu total annuel"],
            ["Bénéfice net", "=Calculs!B8", "Bénéfice net après déduction des coûts"],
            ["ROI", "=Calculs!B8 / Configuration!B14", "Retour sur investissement"],
            ["Temps retour", "=Configuration!B17 / Calculs!B8", "Durée d'amortissement"]
        ]
        
        # Écrire les données dans les feuilles
        config_sheet.update('A1', config_data)
        results_sheet.update('A1', results_data)
        calculations_sheet.update('A1', calculations_data)
        simulation_sheet.update('A1', simulation_formulas)
        
        # Ajouter des notes explicatives
        config_sheet.insert_note('A1', 'Configuration du projet - Ne pas modifier')
        results_sheet.insert_note('A1', 'Résultats de la simulation - Ne pas modifier')
        calculations_sheet.insert_note('A1', 'Calculs intermédiaires - Ne pas modifier')
        simulation_sheet.insert_note('A1', 'Feuille de simulation - Modifiez les valeurs ici')
        
        # Ajouter des formats conditionnels pour les calculs
        calculations_sheet.format('A1:B100', {
            'backgroundColor': {'red': 0.9, 'green': 0.9, 'blue': 0.9},
            'textFormat': {'bold': True}
        })
        
        st.success("Données exportées avec succès vers Google Sheets")
        st.info(f"Ouvrir le fichier : {spreadsheet.url}")
        st.info("""
        Structure des feuilles :
        - "Configuration" : Paramètres de base (ne pas modifier)
        - "Résultats" : Résultats de la simulation (ne pas modifier)
        - "Calculs" : Calculs intermédiaires (ne pas modifier)
        - "Simulation" : Feuille de travail (à modifier pour faire vos propres simulations)
        
        Pour faire une nouvelle simulation :
        1. Copiez la feuille "Simulation"
        2. Modifiez les valeurs dans la nouvelle feuille
        3. Les formules se mettront à jour automatiquement
        4. Les résultats seront mis à jour en temps réel
        """)
        
        return spreadsheet.url
        
    except Exception as e:
        st.error(f"Erreur lors de l'export vers Google Sheets: {str(e)}")
        return None

def export_to_excel(data, filename="simulation_rhum.xlsx"):
    """
    Exporte les données au format Excel (.xlsx) avec une structure claire et formatée
    
    Args:
        data (dict): Données à exporter
        filename (str): Nom du fichier Excel
        
    Returns:
        dict: Données exportées
    """
    try:
        # Créer un nouveau workbook Excel
        wb = Workbook()
        
        # 1. Feuille de Configuration
        config_ws = wb.active
        config_ws.title = "Configuration"
        
        # Style pour les titres
        title_font = Font(bold=True, size=14)
        header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        
        # Ajouter les métadonnées
        config_ws.append(["Configuration Générale"])
        config_ws['A1'].font = title_font
        config_ws.append(["ID du Projet", RHUMA_ID])
        config_ws.append(["Label du Projet", RHUMA_LABEL])
        config_ws.append(["Version", os.getenv('RHUMA_VERSION', "1.0.0")])
        config_ws.append(["Timestamp", datetime.now().isoformat()])
        
        # Espacement
        config_ws.append([])
        
        # Ajouter les paramètres de simulation
        config_ws.append(["Paramètres de Simulation"])
        config_ws['A7'].font = title_font
        
        # Ajouter les paramètres avec style
        for idx, (key, value) in enumerate(RHUMA["configuration"].items(), start=8):
            cell = config_ws.cell(row=idx, column=1, value=key)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="right")
            config_ws.cell(row=idx, column=2, value=value)
        
        # 2. Feuille de Résultats
        results_ws = wb.create_sheet("Résultats")
        
        # Préparer les données de résultats
        results_data = {
            "Production PV (serre)": data.get("production_pv", 0),
            "Production PV (au sol)": data.get("production_au_sol", 0),
            "Production totale": data.get("production_pv", 0) + data.get("production_au_sol", 0),
            "Autoconsommation": data.get("autoconsommation", 0),
            "Revente": data.get("revente", 0),
            "Revenu PV": data.get("revenu_pv", 0),
            "Revenu Rhum": data.get("revenu_rhum", 0),
            "Revenu total": data.get("revenu_pv", 0) + data.get("revenu_rhum", 0),
            "Coût PV": data.get("cout_pv", 0),
            "Coût serre": data.get("cout_serre", 0),
            "Coût total": data.get("cout_total", 0),
            "Bénéfice net": data.get("benefice_net", 0),
            "ROI": data.get("roi", 0),
            "Temps retour": data.get("temps_retour", 0)
        }
        
        # Ajouter les données
        results_ws.append(["Résultats de la Simulation"])
        results_ws['A1'].font = title_font
        
        for row in results_data.items():
            results_ws.append(row)
            results_ws.cell(row=results_ws.max_row, column=1).fill = header_fill
            results_ws.cell(row=results_ws.max_row, column=1).alignment = Alignment(horizontal="right")
        
        # 3. Feuille de Production Mensuelle
        monthly_ws = wb.create_sheet("Production Mensuelle")
        
        # Préparer les données de production mensuelle
        monthly_data = list(data.get("monthly_production", {}).items())
        monthly_ws.append(["Mois", "Production (kWh)"])
        
        for month, production in monthly_data:
            monthly_ws.append([month, production])
        
        # 4. Feuille de Scénarios
        scenarios_ws = wb.create_sheet("Scénarios")
        
        # Préparer les données de scénarios
        scenarios_data = data.get("scenarios", [])
        if scenarios_data:
            scenarios_ws.append(["Scénario", "Description", "Valeur"])
            for scenario in scenarios_data:
                scenarios_ws.append([
                    scenario.get("nom", ""),
                    scenario.get("description", ""),
                    scenario.get("valeur", "")
                ])
        
        # Sauvegarder le fichier Excel
        wb.save(filename)
        
        # Télécharger le fichier
        with open(filename, 'rb') as f:
            st.download_button(
                label="Télécharger l'export Excel",
                data=f,
                file_name=filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        
        st.success("Données exportées avec succès au format Excel")
        
        return data
        
    except Exception as e:
        st.error(f"Erreur lors de l'export Excel: {str(e)}")
        return None

# Sidebar - Paramètres du projet
st.sidebar.header("Paramètres d'Entrée")

# Coûts de construction
st.sidebar.subheader("Coûts de Construction")

# Coûts PV
st.sidebar.markdown("### Coûts PV")
cout_fixe = st.sidebar.number_input(
    "Coût système PV fixe (€/kWc)",
    min_value=0,
    value=int(os.getenv('RHUMA_COUT_FIXE', 1000)),  # Prix moyen d'un système PV fixe en 2024
    step=100,
    help="Coût d'installation d'un système PV fixe par kWc"
)

# Coûts tracking
cout_tracking = st.sidebar.number_input(
    "Coût supplémentaire tracking (€/kWc)",
    min_value=0,
    value=int(os.getenv('RHUMA_COUT_TRACKING', 250)),  # Prix moyen du système tracking en 2024
    step=50,
    help="Coût supplémentaire pour le système de tracking par kWc"
)

# Coûts de construction
cout_construction = st.sidebar.number_input(
    "Coût construction serre (€/m²)",
    min_value=0,
    value=int(os.getenv('RHUMA_COUT_CONSTRUCTION', 150)),  # Prix moyen d'une serre standard en 2024
    step=50,
    help="Coût de construction de la serre par m². Pour une serre standard, les prix varient généralement entre 100 et 200€/m² selon les équipements."
)

# Coûts annuels
st.sidebar.markdown("### Coûts Annuels")
cout_maintenance = st.sidebar.number_input(
    "Coût maintenance annuel (€/kWc)",
    min_value=0,
    value=int(os.getenv('RHUMA_COUT_MAINTENANCE', 50)),  # Prix moyen de maintenance en 2024
    step=10,
    help="Coût annuel de maintenance par kWc"
)

cout_assurance = st.sidebar.number_input(
    "Coût assurance annuel (€/kWc)",
    min_value=0,
    value=int(os.getenv('RHUMA_COUT_ASSURANCE', 20)),  # Prix moyen d'assurance en 2024
    step=5,
    help="Coût annuel d'assurance par kWc"
)

cout_production = st.sidebar.number_input(
    "Coût production annuel (€/kWc)",
    min_value=0,
    value=float(os.getenv('RHUMA_COUT_PRODUCTION', 30.0)),
    step=1.0,
    help="Coût annuel de production de la canne à sucre et distillation"
)

# Calcul des coûts totaux
def calculate_total_costs(puissance_pv, surface_serre):
    """
    Calcule les coûts totaux du projet
    """
    # Coûts initiaux
    cout_pv = puissance_pv * (cout_fixe + cout_tracking)
    cout_serre = surface_serre * cout_construction
    
    # Coûts annuels
    couts_annuels = {
        'maintenance': puissance_pv * cout_maintenance,
        'assurance': puissance_pv * cout_assurance,
        'production': puissance_pv * cout_production
    }
    
    return {
        'cout_pv': cout_pv,
        'cout_serre': cout_serre,
        'cout_total': cout_pv + cout_serre,
        'couts_annuels': couts_annuels
    }

# Paramètres financiers
cost_col1, cost_col2 = st.sidebar.columns(2)

with cost_col1:
    # Tarifs et revenus
    st.write("### Tarifs et Revenus")
    
    # Tarifs EDF
    st.write("### Tarifs EDF")
    tarif_s24 = st.number_input(
        "Tarif S24 (€/kWh)",
        min_value=0.0,
        value=float(os.getenv('RHUMA_TARIF_S24', 0.13)),
        step=0.0001,
        help="Tarif de rachat S24 pour la Corse"
    )
    
    tarif_heures_creuses = st.number_input(
        "Tarif Heures Creuses (€/kWh)",
        min_value=0.0,
        value=float(os.getenv('RHUMA_TARIF_HEURES_CREUSES', 0.15)),
        step=0.0001,
        help="Tarif d'achat pour l'autoconsommation collective"
    )
    
    # Autoconsommation
    st.write("### Autoconsommation")
    autoconsommation_fixe = st.number_input(
        "Autoconsommation système fixe (kWh)",
        min_value=0.0,
        value=float(os.getenv('RHUMA_AUTOCONSOOMMATION_FIXE', 100000.0)),
        step=1000.0,
        help="Quantité d'énergie autoconsommée par an"
    )
    
    autoconsommation_tracking = st.number_input(
        "Autoconsommation système tracking (kWh)",
        min_value=0.0,
        value=float(os.getenv('RHUMA_AUTOCONSOOMMATION_TRACKING', 120000.0)),
        step=1000.0,
        help="Quantité d'énergie autoconsommée par an"
    )

with cost_col2:
    st.write("### Coûts de Production")
    prix_rhum = st.number_input(
        "Prix du rhum (€/L)",
        min_value=0.0,
        value=float(os.getenv('RHUMA_PRIX_RHUM', 20.0)),
        step=1.0,
        help="Prix de vente du rhum"
    )

# Paramètres techniques
tech_col1, tech_col2 = st.sidebar.columns(2)

with tech_col1:
    st.subheader("⚙️ Paramètres Techniques")
    
    # Paramètres de production
    st.write("### Production PV")
    puissance_pv = st.number_input(
        "Puissance installée (kWc)",
        min_value=0.0,
        value=int(os.getenv('RHUMA_PUISSANCE_PV', 500.0)),
        step=50.0,
        help="Puissance totale du système PV"
    )
    
    losses_pv = st.number_input(
        "Pertes PV (%)",
        min_value=0.0,
        max_value=100.0,
        value=int(os.getenv('RHUMA_PERTES_PV', 10.0)),
        step=1.0,
        help="Pertes techniques du système PV"
    )

with tech_col2:
    st.write("### Tracking")
    pertes_tracking = st.number_input(
        "Pertes de tracking (%)",
        min_value=0.0,
        max_value=100.0,
        value=int(os.getenv('RHUMA_PERTES_TRACKING', 5.0)),
        step=1.0,
        help="Pertes liées à l'absence de trackers solaires"
    )
    
    precision_tracking = st.number_input(
        "Précision tracking (°)",
        min_value=0.0,
        max_value=5.0,
        value=float(os.getenv('RHUMA_PRECISION_TRACKING', 0.2)),
        step=0.1,
        help="Précision du système de tracking"
    )

# Paramètres économiques
econ_col1, econ_col2 = st.sidebar.columns(2)

with econ_col1:
    st.subheader("🏦 Paramètres Économiques")
    
    taux_interet = st.number_input(
        "Taux d'intérêt annuel (%)",
        min_value=0.0,
        max_value=100.0,
        value=float(os.getenv('RHUMA_TAUX_INTERET', 3.0)),
        step=0.1,
        help="Taux d'intérêt annuel pour le calcul du ROI"
    )
    
    duree_amortissement = st.number_input(
        "Durée d'amortissement (ans)",
        min_value=1,
        max_value=30,
        value=int(os.getenv('RHUMA_DUREE_AMORTISSEMENT', 20)),
        step=1,
        help="Durée sur laquelle l'investissement est amorti"
    )

with econ_col2:
    cout_production = st.number_input(
        "Coût production annuel (€/kWc)",
        min_value=0.0,
        value=float(os.getenv('RHUMA_COUT_PRODUCTION', 30.0)),
        step=1.0,
        help="Coût annuel de production de la canne à sucre et distillation"
    )
    
    prix_alcool = st.number_input(
        "Prix de l'alcool (€/L)",
        min_value=0.0,
        value=float(os.getenv('RHUMA_PRIX_ALCOOL', 20.0)),
        step=1.0,
        help="Prix de vente de l'alcool"
    )

# 1. Surface et Rendement
if 'surface_canne' not in st.session_state:
    st.session_state.surface_canne = SURFACE_CANNE_DEFAULT
if 'rendement_canne' not in st.session_state:
    st.session_state.rendement_canne = RENDEMENT_CANNE_DEFAULT
if 'teneur_sucre' not in st.session_state:
    st.session_state.teneur_sucre = TENEUR_SUCRE_DEFAULT

surface_canne = st.sidebar.number_input(state_manager.get_i18n('surface_canne'), 
                                      SURFACE_CANNE_MIN, SURFACE_CANNE_MAX, 
                                      st.session_state.surface_canne, 
                                      key='surface_canne')

rendement_canne = st.sidebar.slider(state_manager.get_i18n('rendement_canne'), 
                                  RENDEMENT_CANNE_MIN, RENDEMENT_CANNE_MAX, 
                                  st.session_state.rendement_canne, 
                                  key='rendement_canne')

teneur_sucre = st.sidebar.slider(state_manager.get_i18n('teneur_sucre'), 
                              TENEUR_SUCRE_MIN, TENEUR_SUCRE_MAX, 
                              st.session_state.teneur_sucre, 
                              key='teneur_sucre')

# 2. Extraction et Distillation
if 'efficacite_extraction' not in st.session_state:
    st.session_state.efficacite_extraction = EFFICACITE_EXTRACTION_DEFAULT
if 'efficacite_distillation' not in st.session_state:
    st.session_state.efficacite_distillation = EFFICACITE_DISTILLATION_DEFAULT

efficacite_extraction = st.sidebar.slider(state_manager.get_i18n('efficacite_extraction'), 
                                      EFFICACITE_EXTRACTION_MIN, 
                                      EFFICACITE_EXTRACTION_MAX, 
                                      st.session_state.efficacite_extraction, 
                                      key='efficacite_extraction')

efficacite_distillation = st.sidebar.slider(state_manager.get_i18n('efficacite_distillation'), 
                                        EFFICACITE_DISTILLATION_MIN, 
                                        EFFICACITE_DISTILLATION_MAX, 
                                        st.session_state.efficacite_distillation, 
                                        key='efficacite_distillation')

# 3. Énergie PV
if 'pv_serre' not in st.session_state:
    st.session_state.pv_serre = PV_SERRE_DEFAULT
if 'pv_sol' not in st.session_state:
    st.session_state.pv_sol = PV_SOL_DEFAULT

puissance_pv = st.sidebar.number_input(state_manager.get_i18n('pv_serre'), 
                                    100, PV_SERRE_MAX, 
                                    st.session_state.pv_serre, 
                                    key='pv_serre',
                                    help=f"Limite légale : {LIMITE_PUISSANCE_S24} kWc pour bénéficier du tarif S24")
tarif_s24 = st.sidebar.number_input(state_manager.get_i18n('tarif_s24'), 
                                 TARIF_S24_MIN, TARIF_S24_MAX, 
                                 st.session_state.tarif_s24, 
                                 key='tarif_s24',
                                 help="Tarif de rachat S24 pour la Corse")
tarif_tva = st.sidebar.number_input(state_manager.get_i18n('tva'), 
                                 TVA_MIN, TVA_MAX, 
                                 st.session_state.tva, 
                                 key='tva')
tarif_taxes = st.sidebar.number_input(state_manager.get_i18n('tarif_taxes'), 
                                  0, 100, 
                                  0)
peak_efficiency = st.sidebar.slider(state_manager.get_i18n('peak_efficiency'), 
                                 15, 25, 
                                 20)

# 4. Énergie solaire et heures d'ensoleillement
losses_pv = st.sidebar.slider(state_manager.get_i18n('losses_pv'), 
                            0, 100, 
                            12, 
                            help="12% arrondi de 11.78 selon PVGIS")
autoconsommation = st.sidebar.slider(state_manager.get_i18n('autoconsommation'), 
                                  0, 100, 
                                  30, 
                                  help="Partie de l'électricité utilisée pour la production de rhum")
pertes_tracking = st.sidebar.slider(state_manager.get_i18n('pertes_tracking'), 
                                 0, 100, 
                                 15,
                                 help="Pertes liées à l'absence de trackers solaires")
prix_alcool = st.sidebar.number_input(state_manager.get_i18n('prix_alcool'), 
                                    0.0, 100.0, 
                                    20.0,
                                 help="L'alcool n'est qu'une partie du prix du rhum, s'ajoute d'autres coûts et taxes")

# 5. Énergie PV au sol
if 'pv_sol' not in st.session_state:
    st.session_state.pv_sol = PV_SOL_DEFAULT

capacite_au_sol = st.sidebar.number_input(state_manager.get_i18n('pv_sol'), 
                                          100, PV_SOL_MAX, 
                                          st.session_state.pv_sol, 
                                          key='pv_sol',
                                          help="Production supplémentaire grâce aux panneaux au sol")

# 6. Limites réglementaires
st.sidebar.header(state_manager.get_i18n('limites_reglementaires'))
with st.sidebar.expander(state_manager.get_i18n('details_limites')):
    st.write("- **Tarif S24** :")
    st.write(f"  - Limite de puissance : {LIMITE_PUISSANCE_S24} kWc")
    st.write(f"  - Limite de production : {LIMITE_PRODUCTION_S24/1000:.0f} MWh/an")
    st.write(f"  - Tarif jusqu'à la limite : {tarif_s24}€/kWh")
    st.write(f"  - Tarif au-delà : {TARIF_S24_DEPASSEMENT}€/kWh")
    st.write("- **Production annuelle** :")
    st.write(f"  - Base : {HEURES_PLEIN_SOLEIL} heures équivalent pleine puissance")
    st.write("  - Au-delà : tarif dégradé")

# 7. Tarification clients
st.sidebar.header(state_manager.get_i18n('tarification_clients'))
with st.sidebar.expander(state_manager.get_i18n('tarification_clients_details')):
    st.write("- **Électricité Verte** :")
    st.write("  - Tarif : Des heures toujours creuses !")
    st.write("  - Application : 24h/24h")
    st.write("- **Rhum** :")
    st.write(f"  - Prix : {prix_alcool}€/L")
    st.write("  - Production : 100% locale")

# Production mensuelle selon PVGIS
monthly_production = {
    "janvier": 106.44,
    "février": 118.02,
    "mars": 152.77,
    "avril": 166.63,
    "mai": 184.76,
    "juin": 193.52,
    "juillet": 214.82,
    "août": 204.68,
    "septembre": 167.58,
    "octobre": 142.96,
    "novembre": 100.93,
    "décembre": 103.13
}

# Calcul du total annuel
total_annuel = sum(monthly_production.values())

# Calcul de la production PV mensuelle en fonction de la puissance installée
# Utilisation de la formule PVGIS : Production PV = puissance * irradiation * (1 - pertes/100)
monthly_pv_production = {}
monthly_pv_production_ideal = {}
for month, irradiation in monthly_production.items():
    # Production PV idéale (sans pertes de tracking)
    monthly_pv_production_ideal[month] = puissance_pv * irradiation * (1 - losses_pv / 100)
    # Production PV avec pertes de tracking
    monthly_pv_production[month] = monthly_pv_production_ideal[month] * (1 - pertes_tracking / 100)

# Calcul de la production PV au sol
monthly_au_sol_production = {}
monthly_au_sol_production_ideal = {}
for month, irradiation in monthly_production.items():
    # Production au sol idéale (sans pertes de tracking)
    monthly_au_sol_production_ideal[month] = capacite_au_sol * irradiation * (1 - losses_pv / 100)
    # Production au sol avec pertes de tracking
    monthly_au_sol_production[month] = monthly_au_sol_production_ideal[month] * (1 - pertes_tracking / 100)

# Calcul des productions totales
production_pv = sum(monthly_pv_production.values())
production_pv_ideal = sum(monthly_pv_production_ideal.values())
production_au_sol = sum(monthly_au_sol_production.values())
production_au_sol_ideal = sum(monthly_au_sol_production_ideal.values())
production_totale = production_pv + production_au_sol
production_totale_ideal = production_pv_ideal + production_au_sol_ideal

# Calcul des productions vendues
# Calcul de l'autoconsommation en kWh
autoconsommation_kWh = production_pv * (autoconsommation / 100)

# Calcul de la production vendue
production_vendue = production_pv - autoconsommation_kWh
production_vendue_ideal = production_pv_ideal - autoconsommation_kWh  # Même autoconsommation pour l'idéal

# Calcul du tarif collectif
tarif_collectif = 0.1772 * (1 + tarif_tva/100) * (1 + tarif_taxes/100)

# Modification du calcul du revenu PV
def calcul_tarif_production(production_kwh):
    """Calcule le revenu en tenant compte du seuil des 1600h à pleine puissance"""
    if production_kwh <= LIMITE_PRODUCTION_S24:
        return production_kwh * 0.1772
    else:
        return (LIMITE_PRODUCTION_S24 * 0.1772 + 
                (production_kwh - LIMITE_PRODUCTION_S24) * 0.05)

# Remplacement du calcul simple par le nouveau calcul
revenu_pv = calcul_tarif_production(production_vendue)
revenu_pv_ideal = calcul_tarif_production(production_vendue_ideal)

# Calcul des CA idéaux et réels
chiffre_affaires_collectif = production_au_sol * tarif_collectif
chiffre_affaires_collectif_ideal = production_au_sol_ideal * tarif_collectif
chiffre_affaires_total = revenu_pv + chiffre_affaires_collectif
chiffre_affaires_total_ideal = revenu_pv_ideal + chiffre_affaires_collectif_ideal
delta_ca = chiffre_affaires_total_ideal - chiffre_affaires_total

# Création du tableau
data = {
    "Mois": list(monthly_production.keys()) + ["Total annuel"],
    "Irradiation (kWh/m²)": list(monthly_production.values()) + [total_annuel],
    "Production serre idéale (MWh)": [x/1000 for x in list(monthly_pv_production_ideal.values())] + [production_pv_ideal/1000],
    "Production serre (MWh)": [x/1000 for x in list(monthly_pv_production.values())] + [production_pv/1000]
}

df = pd.DataFrame(data)

# Création du tableau de production au sol
data_au_sol = {
    "Mois": list(monthly_production.keys()) + ["Total annuel"],
    "Production au sol idéale (MWh)": [x/1000 for x in list(monthly_au_sol_production_ideal.values())] + [production_au_sol_ideal/1000],
    "Production au sol (MWh)": [x/1000 for x in list(monthly_au_sol_production.values())] + [production_au_sol/1000]
}

df_au_sol = pd.DataFrame(data_au_sol)

# Affichage du tableau
st.write("\n## Production électrique de la serre")
st.dataframe(df)

# Affichage du tableau de production au sol
st.write("\n## Production électrique au sol")
st.dataframe(df_au_sol)

# Affichage du total annuel
st.write(f"\n### Total Annuel")
st.write(f"- Irradiation totale : {total_annuel:.2f} kWh/m²")
st.write(f"- Production serre idéale : {production_pv_ideal/1000:.1f} MWh")
st.write(f"- Production serre réelle : {production_pv/1000:.1f} MWh")

# Calculs de production
def calcul_production(surface, rendement, sucre, extraction, distillation):
    canne_kg = (surface / 10000) * rendement * 1000  # kg
    sucre_kg = canne_kg * (sucre / 100) * (extraction / 100)
    alcool_l = sucre_kg * 0.51 * (distillation / 100)
    return canne_kg, sucre_kg, alcool_l

canne, sucre, alcool = calcul_production(surface_canne, rendement_canne, teneur_sucre, 
                                        efficacite_extraction, efficacite_distillation)

# Calcul revenus
revenu_rhum = alcool * prix_alcool  # Prix de l'alcool paramétrable

# Barre de progression pour les calculs
with st.spinner("🔄 Calculs en cours..."):
    # Affichage des résultats

    # Organisation des résultats dans des sections collapsibles
    with st.expander("📊 Résultats de la Production de Rhum"):
        st.write("\n## Production de Rhum")
        col1, col2, col3 = st.columns(3)
        col1.metric("📦 Production Canne", f"{canne/1000:.1f} t")
        col2.metric("🍬 Sucre Extrait", f"{sucre/1000:.1f} t")
        col3.metric("🥃 Alcool Pur", f"{alcool:.0f} L")

    with st.expander("⚡ Résultats de la Production d'Énergie"):
        st.write("\n## Production d'Énergie")
        col1, col2, col3 = st.columns(3)
        col1.metric("⚡ Puissance PV (serre) installée", f"{puissance_pv:.0f} kWc")
        col2.metric("⚡ Production serre idéale", f"{production_pv_ideal/1000:.1f} MWh/an")
        col3.metric("⚡ Production serre réelle", f"{production_pv/1000:.1f} MWh/an")

    # Détails de l'autoconsommation
    st.write("\n## Détails de l'Autoconsommation")
    col1, col2, col3 = st.columns(3)
    col1.metric("⚡ Autoconsommation", f"{autoconsommation}%")
    col2.metric("⚡ Électricité autoconsommée", f"{autoconsommation_kWh/1000:.1f} MWh/an")
    col3.metric("⚡ Électricité vendue", f"{production_vendue/1000:.1f} MWh/an")

    # Auto-consommation collective
    st.write("\n## Auto-consommation collective")
    col1, col2, col3 = st.columns(3)
    col1.metric("⚡ Capacité collective", f"{capacite_au_sol:.0f} kWc")
    col2.metric("⚡ Production au sol idéale", f"{production_au_sol_ideal/1000:.1f} MWh/an")
    col3.metric("⚡ Production au sol réelle", f"{production_au_sol/1000:.1f} MWh/an")

    # Résumé du CA
    st.write("\n## Résumé du CA")
    col1, col2, col3 = st.columns(3)
    col1.metric("📊 CA total", f"{chiffre_affaires_total:.0f} €/an")
    col2.metric("📊 CA idéal", f"{chiffre_affaires_total_ideal:.0f} €/an")
    col3.metric("📊 Delta CA", f"{delta_ca:.0f} €/an")

    # Ajout d'espaces verticaux pour une meilleure lisibilité
    st.write("")
    st.write("")

    # Graphiques
    fig, ax = plt.subplots()
    # Calcul des surfaces en m² pour 1 hectare (10000 m²)
    surface_totale = 10000
    surface_locaux = 1000  # 10% de l'hectare
    surface_canne = surface_canne  # Surface dédiée à la canne
    surface_panneaux = surface_totale - surface_locaux - surface_canne  # Reste pour les panneaux

    # Validation visuelle pour les entrées utilisateur
    if surface_canne + surface_locaux > surface_totale:
        st.error("⚠️ La surface totale dépasse 1 hectare. Veuillez ajuster les paramètres.")

    # Ajout d'icônes et de couleurs dans les graphiques
    # Création du graphique des surfaces
    ax.pie([surface_canne, surface_panneaux, surface_locaux], 
        labels=["Canne à sucre", "Panneaux PV", "Locaux"], 
        colors=["#4CAF50", "#FFC107", "#9E9E9E"],
        autopct='%1.1f%%', startangle=90)
    ax.set_title("Répartition dans la serre", color="#4CAF50")

    # Création du graphique des sources de CA
    fig2, ax2 = plt.subplots()
    sources_ca = [
        "Rhum", "PV (vente)", "PV (au sol)", "PV (idéal)"]
    values_ca = [
        revenu_rhum,
        revenu_pv,
        chiffre_affaires_collectif,
        chiffre_affaires_total_ideal - chiffre_affaires_total
    ]

    # Création du graphique en camembert
    ax2.pie(values_ca, 
            labels=sources_ca, 
            colors=["#4CAF50", "#FFC107", "#9E9E9E", "#607D8B"],
            autopct='%1.1f%%', startangle=90)
    ax2.set_title("Répartition des sources de CA")

    # Affichage des graphiques
    st.pyplot(fig)
    # Ajout d'espaces verticaux pour une meilleure lisibilité
    st.write("")
    st.write("")
    st.pyplot(fig2)

    # Dans votre script Streamlit principal, ajoutez ceci après vos sections existantes
    tracking_optimization_section(production_pv_ideal)
    tracking_comparison_section(production_pv_ideal)
    financial_simulation_section()
    pvgis_analysis_section(RHUMA)

    # Détails techniques
    with st.expander("📊 Détails des Calculs"):
        st.write("### 🏗️ Bâtiment (Serre)")
        st.write(f"- Surface totale : {surface_totale} m² (1 ha)")
        st.write(f"- Surface locaux : {surface_locaux} m² (10%)")
        st.write(f"- Surface panneaux en toiture : {surface_panneaux} m²")
        st.write(f"- Surface canne : {surface_canne} m²")

        st.write("\n### 🍯 Production de Rhum")
        st.write(f"- Rendement canne : {rendement_canne} t/ha")
        st.write(f"- Teneur en sucre : {teneur_sucre}%")
        st.write(f"- Efficacité extraction : {efficacite_extraction}%")
        st.write(f"- Efficacité distillation : {efficacite_distillation}%")
        st.write(f"- Kg de canne/L alcool : {canne/alcool:.1f} kg/L")
        st.write(f"- Prix de l'alcool : {prix_alcool}€/L")

        st.write("\n### 🌞 Énergie Solaire")
        st.write(f"- Puissance PV (serre) : {puissance_pv} kWc")
        st.write(f"- Capacité PV (au sol) : {capacite_au_sol} kWc")
        st.write(f"- Pertes PV : {losses_pv}%")
        st.write(f"- Pertes de tracking : {pertes_tracking}%")
        st.write(f"- Efficacité panneaux : {peak_efficiency}%")
        st.write(f"- Tarif S24 : {0.1772}€/kWh")
        st.write(f"- TVA : {tarif_tva}%")
        st.write(f"- Taxes : {tarif_taxes}%")
        st.write(f"- Tarif collectif : {tarif_collectif:.3f}€/kWh")

        st.write("\n### 📊 Production Électrique")
        st.write(f"- Production serre idéale : {production_pv_ideal/1000:.1f} MWh")
        st.write(f"- Production serre réelle : {production_pv/1000:.1f} MWh")
        st.write(f"- Production au sol idéale : {production_au_sol_ideal/1000:.1f} MWh")
        st.write(f"- Production au sol réelle : {production_au_sol/1000:.1f} MWh")
        st.write(f"- Production totale idéale : {(production_pv_ideal + production_au_sol_ideal)/1000:.1f} MWh")
        st.write(f"- Production totale réelle : {(production_pv + production_au_sol)/1000:.1f} MWh")
        st.write(f"- Autoconsommation (%) : {autoconsommation}%")
        st.write(f"- Autoconsommation (MWh) : {autoconsommation_kWh/1000:.1f} MWh")

        st.write("\n### 💰 Revenus")
        st.write(f"- Revenu Rhum : {revenu_rhum:.0f}€/an")
        st.write(f"- Revenu PV (vente) : {revenu_pv:.0f}€/an")
        st.write(f"- CA collectif : {chiffre_affaires_collectif:.0f}€/an")
        st.write(f"- CA collectif idéal : {chiffre_affaires_collectif_ideal:.0f}€/an")
        st.write(f"- CA total : {chiffre_affaires_total:.0f}€/an")
        st.write(f"- CA total idéal : {chiffre_affaires_total_ideal:.0f}€/an")
        st.write(f"- Delta CA : {delta_ca:.0f}€/an")

# Export des résultats
st.header("Exporter les Résultats")

# Bouton d'export complet
if st.button("💾 Exporter tout"):    
    # Préparer les données pour l'export
    data = {
        "production_pv": production_pv,
        "production_au_sol": production_au_sol,
        "autoconsommation": autoconsommation,
        "revente": revente,
        "revenu_pv": revenu_pv,
        "revenu_rhum": revenu_rhum,
        "cout_pv": cout_pv,
        "cout_serre": cout_serre,
        "cout_total": cout_total,
        "benefice_net": benefice_net,
        "roi": roi,
        "temps_retour": temps_retour,
        "monthly_production": monthly_production,
        "scenarios": scenarios
    }
    
    export_all_formats(data)

# Boutons individuels pour chaque format
st.subheader("Exporter un format spécifique")

# Export CSV
if st.button("CSV", key="export_csv"):    
    data = {
        "production_pv": production_pv,
        "production_au_sol": production_au_sol,
        "autoconsommation": autoconsommation,
        "revente": revente,
        "revenu_pv": revenu_pv,
        "revenu_rhum": revenu_rhum,
        "cout_pv": cout_pv,
        "cout_serre": cout_serre,
        "cout_total": cout_total,
        "benefice_net": benefice_net,
        "roi": roi,
        "temps_retour": temps_retour,
        "monthly_production": monthly_production,
        "scenarios": scenarios
    }
    export_to_csv(data)

# Export Excel
if st.button("Excel", key="export_excel"):    
    data = {
        "production_pv": production_pv,
        "production_au_sol": production_au_sol,
        "autoconsommation": autoconsommation,
        "revente": revente,
        "revenu_pv": revenu_pv,
        "revenu_rhum": revenu_rhum,
        "cout_pv": cout_pv,
        "cout_serre": cout_serre,
        "cout_total": cout_total,
        "benefice_net": benefice_net,
        "roi": roi,
        "temps_retour": temps_retour,
        "monthly_production": monthly_production,
        "scenarios": scenarios
    }
    export_to_excel(data)

# Export JSON
if st.button("JSON", key="export_json"):    
    data = {
        "production_pv": production_pv,
        "production_au_sol": production_au_sol,
        "autoconsommation": autoconsommation,
        "revente": revente,
        "revenu_pv": revenu_pv,
        "revenu_rhum": revenu_rhum,
        "cout_pv": cout_pv,
        "cout_serre": cout_serre,
        "cout_total": cout_total,
        "benefice_net": benefice_net,
        "roi": roi,
        "temps_retour": temps_retour,
        "monthly_production": monthly_production,
        "scenarios": scenarios
    }
    export_to_json(data)

# Lien vers le dépôt GitHub
st.markdown("[GitHub Repository](https://github.com/JeanHuguesRobert/Rhuma)")

# Fonction d'export de configuration
def export_config(config_name):
    """
    Exporte les configurations personnalisées dans un fichier .env.config_name
    """
    try:
        # Vérifier le nom de la configuration
        if not config_name or not config_name.replace('_', '').isalnum():
            st.error("Le nom de la configuration doit contenir uniquement des lettres, chiffres et underscores")
            return

        # Créer le contenu du fichier
        config_content = "# Configuration personnalisée\n"
        
        # Ajouter les variables de configuration (exclure les credentials Google Sheets)
        for key, value in os.environ.items():
            if key.startswith('RHUMA_') and not key.startswith('RHUMA_GOOGLE_SHEETS_'):
                config_content += f"{key}={value}\n"

        # Écrire le fichier
        filename = f".env.{config_name}"
        with open(filename, 'w') as f:
            f.write(config_content)

        st.success(f"Configuration exportée avec succès dans {filename}")
        
        # Télécharger le fichier
        with open(filename, 'r') as f:
            st.download_button(
                label="Télécharger la configuration",
                data=f,
                file_name=filename,
                mime="text/plain"
            )

    except Exception as e:
        st.error(f"Erreur lors de l'export de la configuration: {str(e)}")

# Ajouter un séparateur
st.sidebar.markdown("---")

# Export de configuration
config_name = st.sidebar.text_input("Nom de la configuration", "")
if st.sidebar.button("Exporter la configuration"):
    export_config(config_name)

# Paramètres ID
st.sidebar.markdown("---")
st.sidebar.subheader("Paramètres Généraux")

# ID technique
id = st.sidebar.text_input("ID du projet", RHUMA_ID, help="ID technique utilisé pour les fichiers de configuration")
if id != RHUMA_ID:
    if not id.replace('_', '').isalnum():
        st.error("L'ID doit contenir uniquement des lettres, chiffres et underscores")
    else:
        os.environ['RHUMA_ID'] = id

# Label utilisateur
label = st.sidebar.text_input("Label du projet", RHUMA_LABEL, help="Nom affiché dans l'interface utilisateur")
if label != RHUMA_LABEL:
    os.environ['RHUMA_LABEL'] = label

def import_from_json(file_path):
    """
    Importe les données depuis un fichier JSON et gère les versions
    
    Args:
        file_path (str): Chemin vers le fichier JSON
        
    Returns:
        dict: Données importées et adaptées à la version courante
        str: Message d'erreur si l'import échoue
    """
    try:
        # Lire le fichier JSON
        with open(file_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
            
        # Vérifier la version
        file_version = data.get('metadata', {}).get('version', '1.0.0')
        current_version = "1.0.0"  # Version actuelle de l'application
        
        # Comparer les versions
        if file_version != current_version:
            st.warning(f"Attention : le fichier est en version {file_version} alors que l'application est en version {current_version}")
            
            # Adapter les données selon la version
            if file_version == "0.1.0":
                data = adapt_from_v0_1_0(data)
            elif file_version == "0.2.0":
                data = adapt_from_v0_2_0(data)
            # Ajouter d'autres cas de conversion selon les versions
            
        # Vérifier l'intégrité des données
        if not validate_data(data):
            return None, "Les données importées ne sont pas valides"
            
        return data, None
        
    except json.JSONDecodeError:
        return None, "Le fichier n'est pas un JSON valide"
    except Exception as e:
        return None, f"Erreur lors de l'import : {str(e)}"


def adapt_from_v0_1_0(data):
    """
    Adapte les données de la version 0.1.0 à la version actuelle
    """
    # Exemple de conversion
    if 'old_parameter' in data:
        data['new_parameter'] = data.pop('old_parameter')
    
    return data


def adapt_from_v0_2_0(data):
    """
    Adapte les données de la version 0.2.0 à la version actuelle
    """
    # Exemple de conversion
    if 'old_structure' in data:
        data['new_structure'] = convert_structure(data.pop('old_structure'))
    
    return data


def validate_data(data):
    """
    Vérifie l'intégrité des données importées
    """
    required_fields = [
        'metadata',
        'configuration',
        'calculs',
        'scenarios',
        'monthly_production'
    ]
    
    for field in required_fields:
        if field not in data:
            return False
    
    # Vérifier la structure des données
    if not isinstance(data['configuration'], dict):
        return False
    if not isinstance(data['calculs'], dict):
        return False
    
    return True

def export_to_json(data, filename="simulation_results.json"):
    """
    Exporte les résultats de la simulation au format JSON avec une structure optimisée pour les bases de données
    
    Args:
        data (dict): Données à exporter
        filename (str): Nom du fichier JSON
        
    Returns:
        dict: Données exportées
    """
    try:
        # Préparer les données pour l'export
        export_data = {
            "metadata": {
                "id": RHUMA["metadata"]["project_id"],
                "label": RHUMA["metadata"]["project_label"],
                "version": RHUMA["metadata"]["version"],
                "timestamp": RHUMA["metadata"]["timestamp"],
                "export_timestamp": datetime.now().isoformat()
            },
            "configuration": {
                "surface_canne": rhuma('surface_canne'),
                "rendement_canne": rhuma('rendement_canne'),
                "teneur_sucre": rhuma('teneur_sucre'),
                "efficacite_extraction": rhuma('efficacite_extraction'),
                "efficacite_distillation": rhuma('efficacite_distillation'),
                "pv_serre": rhuma('pv_serre'),
                "pv_sol": rhuma('pv_sol'),
                "tarif_s24": rhuma('tarif_s24'),
                "tva": rhuma('tva'),
                "cout_fixe": rhuma('cout_fixe'),
                "cout_tracking": rhuma('cout_tracking'),
                "cout_construction": rhuma('cout_construction'),
                "cout_maintenance": rhuma('cout_maintenance'),
                "cout_assurance": rhuma('cout_assurance'),
                "cout_production": rhuma('cout_production'),
                "tarif_heures_creuses": rhuma('tarif_heures_creuses'),
                "autoconsommation_fixe": rhuma('autoconsommation_fixe'),
                "autoconsommation_tracking": rhuma('autoconsommation_tracking'),
                "prix_rhum": rhuma('prix_rhum'),
                "pertes_pv": rhuma('pertes_pv'),
                "pertes_tracking": rhuma('pertes_tracking'),
                "precision_tracking": rhuma('precision_tracking'),
                "taux_interet": rhuma('taux_interet'),
                "duree_amortissement": rhuma('duree_amortissement')
            },
            "results": {
                "production": {
                    "pv_serre": float(data.get("production_pv", 0)),
                    "pv_sol": float(data.get("production_au_sol", 0)),
                    "total": float(data.get("production_pv", 0) + data.get("production_au_sol", 0)),
                    "autoconsommation": float(data.get("autoconsommation", 0)),
                    "revente": float(data.get("revente", 0))
                },
                "revenus": {
                    "pv": float(data.get("revenu_pv", 0)),
                    "rhum": float(data.get("revenu_rhum", 0)),
                    "total": float(data.get("revenu_pv", 0) + data.get("revenu_rhum", 0))
                },
                "costs": {
                    "initial": {
                        "pv": float(data.get("cout_pv", 0)),
                        "serre": float(data.get("cout_serre", 0)),
                        "total": float(data.get("cout_total", 0))
                    },
                    "annual": {
                        "maintenance": float(data.get("couts_annuels", {}).get("maintenance", 0)),
                        "insurance": float(data.get("couts_annuels", {}).get("assurance", 0)),
                        "production": float(data.get("couts_annuels", {}).get("production", 0))
                    }
                },
                "benefits": {
                    "net": float(data.get("benefice_net", 0)),
                    "roi": float(data.get("roi", 0)),
                    "return_time": float(data.get("temps_retour", 0))
                }
            },
            "monthly_production": {
                "data": [{
                    "month": month,
                    "production": float(production)
                } for month, production in data.get("monthly_production", {}).items()]
            },
            "scenarios": [{
                "id": idx + 1,
                "name": scenario.get("nom", ""),
                "description": scenario.get("description", ""),
                "value": float(scenario.get("valeur", 0)) if scenario.get("valeur") is not None else None
            } for idx, scenario in enumerate(data.get("scenarios", []))]
        }

        # Écrire le fichier JSON
        with open(filename, 'w', encoding='utf-8') as f:
            json.dump(export_data, f, indent=4, ensure_ascii=False)
        
        # Télécharger le fichier
        with open(filename, 'r') as f:
            st.download_button(
                label="Télécharger le fichier JSON",
                data=f.read(),
                file_name=filename,
                mime="application/json"
            )
        
        st.success(f"Données exportées avec succès dans {filename}")
        
        return export_data
        
    except Exception as e:
        st.error(f"Erreur lors de l'export JSON: {str(e)}")
        return None

def export_all_formats(data, filename_prefix="simulation_rhum"):
    """
    Exporte les données dans tous les formats disponibles (JSON, CSV, Excel, .env)
    et les regroupe dans un fichier ZIP unique
    
    Args:
        data (dict): Données à exporter
        filename_prefix (str): Préfixe pour les noms de fichiers
        
    Returns:
        dict: Données exportées
    """
    try:
        # Créer un dossier temporaire pour stocker les fichiers
        temp_dir = tempfile.mkdtemp()
        
        # 1. Export JSON
        json_filename = os.path.join(temp_dir, f"{filename_prefix}.json")
        with open(json_filename, 'w', encoding='utf-8') as f:
            json.dump(export_to_json(data), f, indent=4, ensure_ascii=False)
        
        # 2. Export Excel
        excel_filename = os.path.join(temp_dir, f"{filename_prefix}.xlsx")
        export_to_excel(data, excel_filename)
        
        # 3. Export CSV (dans un dossier séparé)
        csv_dir = os.path.join(temp_dir, "csv")
        os.makedirs(csv_dir, exist_ok=True)
        
        # Configuration
        config_df = pd.DataFrame([
            ["ID du Projet", RHUMA["metadata"]["project_id"]],
            ["Label du Projet", RHUMA["metadata"]["project_label"]],
            ["Version", RHUMA["metadata"]["version"]],
            ["Timestamp", RHUMA["metadata"]["timestamp"]]
        ])
        config_df.to_csv(os.path.join(csv_dir, f"{filename_prefix}_configuration.csv"), index=False, encoding='utf-8')
        
        # Paramètres
        params_df = pd.DataFrame([
            ["Surface canne (m²)", RHUMA["configuration"]["surface_canne"]],
            ["Rendement canne (t/ha)", RHUMA["configuration"]["rendement_canne"]],
            ["Teneur sucre (%)", RHUMA["configuration"]["teneur_sucre"]],
            ["Efficacité extraction (%)", RHUMA["configuration"]["efficacite_extraction"]],
            ["Efficacité distillation (%)", RHUMA["configuration"]["efficacite_distillation"]],
            ["Puissance PV (serre) (kWc)", RHUMA["configuration"]["pv_serre"]],
            ["Puissance PV (au sol) (kWc)", RHUMA["configuration"]["pv_sol"]],
            ["Tarif S24 (€/kWh)", RHUMA["configuration"]["tarif_s24"]],
            ["TVA (%)", RHUMA["configuration"]["tva"]],
            ["Coût système PV fixe (€/kWc)", RHUMA["configuration"]["cout_fixe"]],
            ["Coût système tracking (€/kWc)", RHUMA["configuration"]["cout_tracking"]],
            ["Coût construction serre (€/m²)", RHUMA["configuration"]["cout_construction"]],
            ["Coût maintenance (€/kWc)", RHUMA["configuration"]["cout_maintenance"]],
            ["Coût assurance (€/kWc)", RHUMA["configuration"]["cout_assurance"]],
            ["Coût production (€/kWc)", RHUMA["configuration"]["cout_production"]],
            ["Tarif heures creuses (€/kWh)", RHUMA["configuration"]["tarif_heures_creuses"]],
            ["Autoconsommation fixe (kWh)", RHUMA["configuration"]["autoconsommation_fixe"]],
            ["Autoconsommation tracking (kWh)", RHUMA["configuration"]["autoconsommation_tracking"]],
            ["Prix du rhum (€/L)", RHUMA["configuration"]["prix_rhum"]],
            ["Pertes PV (%)", RHUMA["configuration"]["pertes_pv"]],
            ["Pertes tracking (%)", RHUMA["configuration"]["pertes_tracking"]],
            ["Précision tracking (°)", RHUMA["configuration"]["precision_tracking"]],
            ["Taux d'intérêt (%)", RHUMA["configuration"]["taux_interet"]],
            ["Durée d'amortissement (ans)", RHUMA["configuration"]["duree_amortissement"]]
        ])
        params_df.to_csv(os.path.join(csv_dir, f"{filename_prefix}_parameters.csv"), index=False, encoding='utf-8')
        
        # Résultats
        results_df = pd.DataFrame([
            ["Production PV (serre)", data.get("production_pv", 0)],
            ["Production PV (au sol)", data.get("production_au_sol", 0)],
            ["Production totale", data.get("production_pv", 0) + data.get("production_au_sol", 0)],
            ["Autoconsommation", data.get("autoconsommation", 0)],
            ["Revente", data.get("revente", 0)],
            ["Revenu PV", data.get("revenu_pv", 0)],
            ["Revenu Rhum", data.get("revenu_rhum", 0)],
            ["Revenu total", data.get("revenu_pv", 0) + data.get("revenu_rhum", 0)],
            ["Coût PV", data.get("cout_pv", 0)],
            ["Coût serre", data.get("cout_serre", 0)],
            ["Coût total", data.get("cout_total", 0)],
            ["Bénéfice net", data.get("benefice_net", 0)],
            ["ROI", data.get("roi", 0)],
            ["Temps retour", data.get("temps_retour", 0)]
        ])
        results_df.to_csv(os.path.join(csv_dir, f"{filename_prefix}_results.csv"), index=False, encoding='utf-8')
        
        # Production mensuelle
        monthly_df = pd.DataFrame(list(data.get("monthly_production", {}).items()),
                                columns=["Mois", "Production (kWh)"])
        monthly_df.to_csv(os.path.join(csv_dir, f"{filename_prefix}_monthly_production.csv"), index=False, encoding='utf-8')
        
        # 4. Export .env (sans les secrets)
        env_filename = os.path.join(temp_dir, f"{filename_prefix}.env")
        with open(env_filename, 'w', encoding='utf-8') as f:
            f.write(f"RHUMA_ID={RHUMA['metadata']['project_id']}\n")
            f.write(f"RHUMA_LABEL={RHUMA['metadata']['project_label']}\n")
            f.write(f"RHUMA_VERSION={RHUMA['metadata']['version']}\n")
            for key, value in RHUMA["configuration"].items():
                if isinstance(value, (int, float)):
                    f.write(f"RHUMA_{key.upper()}={value}\n")
        
        # Créer le fichier ZIP
        zip_filename = f"{filename_prefix}_export.zip"
        with zipfile.ZipFile(zip_filename, 'w') as zipf:
            # Ajouter tous les fichiers
            zipf.write(json_filename, os.path.basename(json_filename))
            zipf.write(excel_filename, os.path.basename(excel_filename))
            zipf.write(env_filename, os.path.basename(env_filename))
            
            # Ajouter les fichiers CSV
            for csv_file in os.listdir(csv_dir):
                zipf.write(os.path.join(csv_dir, csv_file), f"csv/{csv_file}")
        
        # Nettoyer les fichiers temporaires
        shutil.rmtree(temp_dir)
        
        # Télécharger le fichier ZIP
        with open(zip_filename, 'rb') as f:
            st.download_button(
                label="Télécharger l'export complet",
                data=f,
                file_name=zip_filename,
                mime="application/zip"
            )
        
        st.success("Données exportées avec succès dans tous les formats")
        
        return data
        
    except Exception as e:
        st.error(f"Erreur lors de l'export complet: {str(e)}")
        return None

# Export des résultats
if st.button("💾 Exporter en CSV"):
    data = {
        "Mois": list(monthly_production.keys()) + ["Total annuel"],
        "Production serre (MWh)": [x/1000 for x in list(monthly_pv_production.values())] + [production_pv/1000],
        "Production collective (MWh)": [x/1000 for x in list(monthly_au_sol_production.values())] + [production_au_sol/1000],
        "CA collectif (€)": [chiffre_affaires_collectif] * 12 + [chiffre_affaires_collectif]
    }
    df_export = pd.DataFrame(data)
    st.download_button("⬇️ Télécharger", df_export.to_csv(index=False), "production_rhum_solaire.csv", "text/csv")

def export_to_csv(data, filename_prefix="simulation_rhum"):
    """
    Exporte les données au format CSV avec une structure claire
    
    Args:
        data (dict): Données à exporter
        filename_prefix (str): Préfixe du nom du fichier
        
    Returns:
        dict: Données exportées
    """
    try:
        # Préparer les données pour l'export
        export_data = {
            "metadata": RHUMA["metadata"],
            "configuration": RHUMA["configuration"],
            "calculs": {
                "production": {
                    "pv_serre": data.get("production_pv", 0),
                    "pv_sol": data.get("production_au_sol", 0),
                    "total": data.get("production_pv", 0) + data.get("production_au_sol", 0),
                    "autoconsommation": data.get("autoconsommation", 0),
                    "revente": data.get("revente", 0)
                },
                "revenus": {
                    "pv": data.get("revenu_pv", 0),
                    "rhum": data.get("revenu_rhum", 0),
                    "total": data.get("revenu_pv", 0) + data.get("revenu_rhum", 0)
                },
                "couts": {
                    "initiaux": {
                        "pv": data.get("cout_pv", 0),
                        "serre": data.get("cout_serre", 0),
                        "total": data.get("cout_total", 0)
                    },
                    "annuels": {
                        "maintenance": data.get("couts_annuels", {}).get("maintenance", 0),
                        "assurance": data.get("couts_annuels", {}).get("assurance", 0),
                        "production": data.get("couts_annuels", {}).get("production", 0)
                    }
                },
                "benefices": {
                    "net": data.get("benefice_net", 0),
                    "roi": data.get("roi", 0),
                    "temps_retour": data.get("temps_retour", 0)
                }
            },
            "scenarios": data.get("scenarios", []),
            "monthly_production": data.get("monthly_production", {})
        }
        
        # Créer les fichiers CSV pour chaque feuille
        filenames = {}
        
        # 1. Configuration
        config_df = pd.DataFrame([
            ["ID du Projet", RHUMA["metadata"]["project_id"]],
            ["Label du Projet", RHUMA["metadata"]["project_label"]],
            ["Version", RHUMA["metadata"]["version"]],
            ["Timestamp", RHUMA["metadata"]["timestamp"]]
        ])
        config_df.to_csv(f"{filename_prefix}_configuration.csv", index=False, encoding='utf-8')
        filenames["configuration"] = f"{filename_prefix}_configuration.csv"
        
        # 2. Paramètres de Simulation
        params_df = pd.DataFrame([
            ["Surface canne (m²)", RHUMA["configuration"]["surface_canne"]],
            ["Rendement canne (t/ha)", RHUMA["configuration"]["rendement_canne"]],
            ["Teneur sucre (%)", RHUMA["configuration"]["teneur_sucre"]],
            ["Efficacité extraction (%)", RHUMA["configuration"]["efficacite_extraction"]],
            ["Efficacité distillation (%)", RHUMA["configuration"]["efficacite_distillation"]],
            ["Puissance PV (serre) (kWc)", RHUMA["configuration"]["pv_serre"]],
            ["Puissance PV (au sol) (kWc)", RHUMA["configuration"]["pv_sol"]],
            ["Tarif S24 (€/kWh)", RHUMA["configuration"]["tarif_s24"]],
            ["TVA (%)", RHUMA["configuration"]["tva"]],
            ["Coût système PV fixe (€/kWc)", RHUMA["configuration"]["cout_fixe"]],
            ["Coût système tracking (€/kWc)", RHUMA["configuration"]["cout_tracking"]],
            ["Coût construction serre (€/m²)", RHUMA["configuration"]["cout_construction"]],
            ["Coût maintenance (€/kWc)", RHUMA["configuration"]["cout_maintenance"]],
            ["Coût assurance (€/kWc)", RHUMA["configuration"]["cout_assurance"]],
            ["Coût production (€/kWc)", RHUMA["configuration"]["cout_production"]],
            ["Tarif heures creuses (€/kWh)", RHUMA["configuration"]["tarif_heures_creuses"]],
            ["Autoconsommation fixe (kWh)", RHUMA["configuration"]["autoconsommation_fixe"]],
            ["Autoconsommation tracking (kWh)", RHUMA["configuration"]["autoconsommation_tracking"]],
            ["Prix du rhum (€/L)", RHUMA["configuration"]["prix_rhum"]],
            ["Pertes PV (%)", RHUMA["configuration"]["pertes_pv"]],
            ["Pertes tracking (%)", RHUMA["configuration"]["pertes_tracking"]],
            ["Précision tracking (°)", RHUMA["configuration"]["precision_tracking"]],
            ["Taux d'intérêt (%)", RHUMA["configuration"]["taux_interet"]],
            ["Durée d'amortissement (ans)", RHUMA["configuration"]["duree_amortissement"]]
        ])
        params_df.to_csv(f"{filename_prefix}_parameters.csv", index=False, encoding='utf-8')
        filenames["parameters"] = f"{filename_prefix}_parameters.csv"
        
        # 3. Résultats
        results_df = pd.DataFrame([
            ["Production PV (serre)", data.get("production_pv", 0)],
            ["Production PV (au sol)", data.get("production_au_sol", 0)],
            ["Production totale", data.get("production_pv", 0) + data.get("production_au_sol", 0)],
            ["Autoconsommation", data.get("autoconsommation", 0)],
            ["Revente", data.get("revente", 0)],
            ["Revenu PV", data.get("revenu_pv", 0)],
            ["Revenu Rhum", data.get("revenu_rhum", 0)],
            ["Revenu total", data.get("revenu_pv", 0) + data.get("revenu_rhum", 0)],
            ["Coût PV", data.get("cout_pv", 0)],
            ["Coût serre", data.get("cout_serre", 0)],
            ["Coût total", data.get("cout_total", 0)],
            ["Bénéfice net", data.get("benefice_net", 0)],
            ["ROI", data.get("roi", 0)],
            ["Temps retour", data.get("temps_retour", 0)]
        ])
        results_df.to_csv(f"{filename_prefix}_results.csv", index=False, encoding='utf-8')
        filenames["results"] = f"{filename_prefix}_results.csv"
        
        # 4. Production mensuelle
        monthly_df = pd.DataFrame(list(data.get("monthly_production", {}).items()),
                                columns=["Mois", "Production (kWh)"])
        monthly_df.to_csv(f"{filename_prefix}_monthly_production.csv", index=False, encoding='utf-8')
        filenames["monthly_production"] = f"{filename_prefix}_monthly_production.csv"
        
        # Créer un fichier ZIP avec tous les fichiers CSV
        zip_filename = f"{filename_prefix}_export.zip"
        with zipfile.ZipFile(zip_filename, 'w') as zipf:
            # Ajouter tous les fichiers
            zipf.write(f"{filename_prefix}_configuration.csv", os.path.basename(f"{filename_prefix}_configuration.csv"))
            zipf.write(f"{filename_prefix}_parameters.csv", os.path.basename(f"{filename_prefix}_parameters.csv"))
            zipf.write(f"{filename_prefix}_results.csv", os.path.basename(f"{filename_prefix}_results.csv"))
            zipf.write(f"{filename_prefix}_monthly_production.csv", os.path.basename(f"{filename_prefix}_monthly_production.csv"))
            os.remove(f"{filename_prefix}_configuration.csv")
            os.remove(f"{filename_prefix}_parameters.csv")
            os.remove(f"{filename_prefix}_results.csv")
            os.remove(f"{filename_prefix}_monthly_production.csv")
        
        # Télécharger le fichier ZIP
        with open(zip_filename, 'rb') as f:
            st.download_button(
                label="Télécharger l'export CSV",
                data=f,
                file_name=zip_filename,
                mime="application/zip"
            )
        
        st.success("Données exportées avec succès au format CSV")
        
        return export_data
        
    except Exception as e:
        st.error(f"Erreur lors de l'export CSV: {str(e)}")
        return None

def export_to_json(data, filename="simulation_results.json"):
    """
    Exporte les résultats de la simulation au format JSON avec une structure optimisée pour les bases de données
    
    Args:
        data (dict): Données à exporter
        filename (str): Nom du fichier JSON
        
    Returns:
        dict: Données exportées
    """
    try:
        # Préparer les données pour l'export
        export_data = {
            "metadata": {
                "id": RHUMA["metadata"]["project_id"],
                "label": RHUMA["metadata"]["project_label"],
                "version": RHUMA["metadata"]["version"],
                "timestamp": RHUMA["metadata"]["timestamp"],
                "export_timestamp": datetime.now().isoformat()
            },
            "configuration": {
                "surface_canne": rhuma('surface_canne'),
                "rendement_canne": rhuma('rendement_canne'),
                "teneur_sucre": rhuma('teneur_sucre'),
                "efficacite_extraction": rhuma('efficacite_extraction'),
                "efficacite_distillation": rhuma('efficacite_distillation'),
                "pv_serre": rhuma('pv_serre'),
                "pv_sol": rhuma('pv_sol'),
                "tarif_s24": rhuma('tarif_s24'),
                "tva": rhuma('tva'),
                "cout_fixe": rhuma('cout_fixe'),
                "cout_tracking": rhuma('cout_tracking'),
                "cout_construction": rhuma('cout_construction'),
                "cout_maintenance": rhuma('cout_maintenance'),
                "cout_assurance": rhuma('cout_assurance'),
                "cout_production": rhuma('cout_production'),
                "tarif_heures_creuses": rhuma('tarif_heures_creuses'),
                "autoconsommation_fixe": rhuma('autoconsommation_fixe'),
                "autoconsommation_tracking": rhuma('autoconsommation_tracking'),
                "prix_rhum": rhuma('prix_rhum'),
                "pertes_pv": rhuma('pertes_pv'),
                "pertes_tracking": rhuma('pertes_tracking'),
                "precision_tracking": rhuma('precision_tracking'),
                "taux_interet": rhuma('taux_interet'),
                "duree_amortissement": rhuma('duree_amortissement')
            },
            "results": {
                "production": {
                    "pv_serre": float(data.get("production_pv", 0)),
                    "pv_sol": float(data.get("production_au_sol", 0)),
                    "total": float(data.get("production_pv", 0) + data.get("production_au_sol", 0)),
                    "autoconsommation": float(data.get("autoconsommation", 0)),
                    "revente": float(data.get("revente", 0))
                },
                "revenus": {
                    "pv": float(data.get("revenu_pv", 0)),
                    "rhum": float(data.get("revenu_rhum", 0)),
                    "total": float(data.get("revenu_pv", 0) + data.get("revenu_rhum", 0))
                },
                "costs": {
                    "initial": {
                        "pv": float(data.get("cout_pv", 0)),
                        "serre": float(data.get("cout_serre", 0)),
                        "total": float(data.get("cout_total", 0))
                    },
                    "annual": {
                        "maintenance": float(data.get("couts_annuels", {}).get("maintenance", 0)),
                        "insurance": float(data.get("couts_annuels", {}).get("assurance", 0)),
                        "production": float(data.get("couts_annuels", {}).get("production", 0))
                    }
                },
                "benefits": {
                    "net": float(data.get("benefice_net", 0)),
                    "roi": float(data.get("roi", 0)),
                    "return_time": float(data.get("temps_retour", 0))
                }
            },
            "monthly_production": {
                "data": [{
                    "month": month,
                    "production": float(production)
                } for month, production in data.get("monthly_production", {}).items()]
            },
            "scenarios": [{
                "id": idx + 1,
                "name": scenario.get("nom", ""),
                "description": scenario.get("description", ""),
                "value": float(scenario.get("valeur", 0)) if scenario.get("valeur") is not None else None
            } for idx, scenario in enumerate(data.get("scenarios", []))]
        }

        # Écrire le fichier JSON
        with open(filename, 'w', encoding='utf-8') as f:
            json.dump(export_data, f, indent=4, ensure_ascii=False)
        
        # Télécharger le fichier
        with open(filename, 'r') as f:
            st.download_button(
                label="Télécharger le fichier JSON",
                data=f.read(),
                file_name=filename,
                mime="application/json"
            )
        
        st.success(f"Données exportées avec succès dans {filename}")
        
        return export_data
        
    except Exception as e:
        st.error(f"Erreur lors de l'export JSON: {str(e)}")
        return None

def export_all_formats(data, filename_prefix="simulation_rhum"):
    """
    Exporte les données dans tous les formats disponibles (JSON, CSV, Excel, .env)
    et les regroupe dans un fichier ZIP unique
    
    Args:
        data (dict): Données à exporter
        filename_prefix (str): Préfixe pour les noms de fichiers
        
    Returns:
        dict: Données exportées
    """
    try:
        # Créer un dossier temporaire pour stocker les fichiers
        temp_dir = tempfile.mkdtemp()
        
        # 1. Export JSON
        json_filename = os.path.join(temp_dir, f"{filename_prefix}.json")
        with open(json_filename, 'w', encoding='utf-8') as f:
            json.dump(export_to_json(data), f, indent=4, ensure_ascii=False)
        
        # 2. Export Excel
        excel_filename = os.path.join(temp_dir, f"{filename_prefix}.xlsx")
        export_to_excel(data, excel_filename)
        
        # 3. Export CSV (dans un dossier séparé)
        csv_dir = os.path.join(temp_dir, "csv")
        os.makedirs(csv_dir, exist_ok=True)
        
        # Configuration
        config_df = pd.DataFrame([
            ["ID du Projet", RHUMA["metadata"]["project_id"]],
            ["Label du Projet", RHUMA["metadata"]["project_label"]],
            ["Version", RHUMA["metadata"]["version"]],
            ["Timestamp", RHUMA["metadata"]["timestamp"]]
        ])
        config_df.to_csv(os.path.join(csv_dir, f"{filename_prefix}_configuration.csv"), index=False, encoding='utf-8')
        
        # Paramètres
        params_df = pd.DataFrame([
            ["Surface canne (m²)", RHUMA["configuration"]["surface_canne"]],
            ["Rendement canne (t/ha)", RHUMA["configuration"]["rendement_canne"]],
            ["Teneur sucre (%)", RHUMA["configuration"]["teneur_sucre"]],
            ["Efficacité extraction (%)", RHUMA["configuration"]["efficacite_extraction"]],
            ["Efficacité distillation (%)", RHUMA["configuration"]["efficacite_distillation"]],
            ["Puissance PV (serre) (kWc)", RHUMA["configuration"]["pv_serre"]],
            ["Puissance PV (au sol) (kWc)", RHUMA["configuration"]["pv_sol"]],
            ["Tarif S24 (€/kWh)", RHUMA["configuration"]["tarif_s24"]],
            ["TVA (%)", RHUMA["configuration"]["tva"]],
            ["Coût système PV fixe (€/kWc)", RHUMA["configuration"]["cout_fixe"]],
            ["Coût système tracking (€/kWc)", RHUMA["configuration"]["cout_tracking"]],
            ["Coût construction serre (€/m²)", RHUMA["configuration"]["cout_construction"]],
            ["Coût maintenance (€/kWc)", RHUMA["configuration"]["cout_maintenance"]],
            ["Coût assurance (€/kWc)", RHUMA["configuration"]["cout_assurance"]],
            ["Coût production (€/kWc)", RHUMA["configuration"]["cout_production"]],
            ["Tarif heures creuses (€/kWh)", RHUMA["configuration"]["tarif_heures_creuses"]],
            ["Autoconsommation fixe (kWh)", RHUMA["configuration"]["autoconsommation_fixe"]],
            ["Autoconsommation tracking (kWh)", RHUMA["configuration"]["autoconsommation_tracking"]],
            ["Prix du rhum (€/L)", RHUMA["configuration"]["prix_rhum"]],
            ["Pertes PV (%)", RHUMA["configuration"]["pertes_pv"]],
            ["Pertes tracking (%)", RHUMA["configuration"]["pertes_tracking"]],
            ["Précision tracking (°)", RHUMA["configuration"]["precision_tracking"]],
            ["Taux d'intérêt (%)", RHUMA["configuration"]["taux_interet"]],
            ["Durée d'amortissement (ans)", RHUMA["configuration"]["duree_amortissement"]]
        ])
        params_df.to_csv(os.path.join(csv_dir, f"{filename_prefix}_parameters.csv"), index=False, encoding='utf-8')
        
        # Résultats
        results_df = pd.DataFrame([
            ["Production PV (serre)", data.get("production_pv", 0)],
            ["Production PV (au sol)", data.get("production_au_sol", 0)],
            ["Production totale", data.get("production_pv", 0) + data.get("production_au_sol", 0)],
            ["Autoconsommation", data.get("autoconsommation", 0)],
            ["Revente", data.get("revente", 0)],
            ["Revenu PV", data.get("revenu_pv", 0)],
            ["Revenu Rhum", data.get("revenu_rhum", 0)],
            ["Revenu total", data.get("revenu_pv", 0) + data.get("revenu_rhum", 0)],
            ["Coût PV", data.get("cout_pv", 0)],
            ["Coût serre", data.get("cout_serre", 0)],
            ["Coût total", data.get("cout_total", 0)],
            ["Bénéfice net", data.get("benefice_net", 0)],
            ["ROI", data.get("roi", 0)],
            ["Temps retour", data.get("temps_retour", 0)]
        ])
        results_df.to_csv(os.path.join(csv_dir, f"{filename_prefix}_results.csv"), index=False, encoding='utf-8')
        
        # Production mensuelle
        monthly_df = pd.DataFrame(list(data.get("monthly_production", {}).items()),
                                columns=["Mois", "Production (kWh)"])
        monthly_df.to_csv(os.path.join(csv_dir, f"{filename_prefix}_monthly_production.csv"), index=False, encoding='utf-8')
        
        # 4. Export .env (sans les secrets)
        env_filename = os.path.join(temp_dir, f"{filename_prefix}.env")
        with open(env_filename, 'w', encoding='utf-8') as f:
            f.write(f"RHUMA_ID={RHUMA['metadata']['project_id']}\n")
            f.write(f"RHUMA_LABEL={RHUMA['metadata']['project_label']}\n")
            f.write(f"RHUMA_VERSION={RHUMA['metadata']['version']}\n")
            for key, value in RHUMA["configuration"].items():
                if isinstance(value, (int, float)):
                    f.write(f"RHUMA_{key.upper()}={value}\n")
        
        # Créer le fichier ZIP
        zip_filename = f"{filename_prefix}_export.zip"
        with zipfile.ZipFile(zip_filename, 'w') as zipf:
            # Ajouter tous les fichiers
            zipf.write(json_filename, os.path.basename(json_filename))
            zipf.write(excel_filename, os.path.basename(excel_filename))
            zipf.write(env_filename, os.path.basename(env_filename))
            
            # Ajouter les fichiers CSV
            for csv_file in os.listdir(csv_dir):
                zipf.write(os.path.join(csv_dir, csv_file), f"csv/{csv_file}")
        
        # Nettoyer les fichiers temporaires
        shutil.rmtree(temp_dir)
        
        # Télécharger le fichier ZIP
        with open(zip_filename, 'rb') as f:
            st.download_button(
                label="Télécharger l'export complet",
                data=f,
                file_name=zip_filename,
                mime="application/zip"
            )
        
        st.success("Données exportées avec succès dans tous les formats")
        
        return data
        
    except Exception as e:
        st.error(f"Erreur lors de l'export complet: {str(e)}")
        return None

# Lien vers le dépôt GitHub
st.markdown("[GitHub Repository](https://github.com/JeanHuguesRobert/Rhuma)")
