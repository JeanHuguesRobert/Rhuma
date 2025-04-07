import os
import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
import gspread
from oauth2client.service_account import ServiceAccountCredentials
import os

def get_google_sheet_client():
    """
    Initialise et retourne un client Google Sheets
    
    Returns:
        gspread.Client: Client Google Sheets configuré
    """
    try:
        # Charger les credentials depuis les variables d'environnement
        creds_json = os.getenv('GOOGLE_SHEETS_CREDENTIALS_JSON')
        if not creds_json:
            return None
            
        # Créer les credentials
        scope = 'https://spreadsheets.google.com/feeds https://www.googleapis.com/auth/drive'
        credentials = ServiceAccountCredentials.from_json_keyfile_dict(
            json.loads(creds_json), scope)
            
        # Créer et retourner le client
        return gspread.authorize(credentials.credentials)
    except Exception as e:
        print(f"Erreur lors de l'initialisation du client Google Sheets: {e}")
        return None

def export_to_google_sheets(data, sheet_name="Simulation Rhuma"):
    """
    Exporte les données vers une nouvelle feuille Google Sheets avec une structure optimisée
    """
    try:
        # Initialiser le client
        client = get_google_sheet_client()
        if not client:
            return "Configuration Google Sheets non valide"
            
        # Créer un nouveau spreadsheet
        spreadsheet = client.create(sheet_name)
        
        # Partager le spreadsheet avec l'utilisateur
        email = os.getenv('RHUMA_GOOGLE_SHEETS_CLIENT_EMAIL', '')  # Default to empty string if not set
        spreadsheet.share(
            email,  # Now guaranteed to be a string
            perm_type='user',
            role='writer'
        )
        
        # Créer les différentes feuilles
        config_sheet = spreadsheet.add_worksheet(title="Configuration", rows="200", cols="20")
        results_sheet = spreadsheet.add_worksheet(title="Résultats", rows="200", cols="20")
        calculations_sheet = spreadsheet.add_worksheet(title="Calculs", rows="200", cols="20")
        simulation_sheet = spreadsheet.add_worksheet(title="Simulation", rows="200", cols="20")
        
        # Préparer les données pour chaque feuille
        config_data = [
            ["Configuration Générale"] + [""] * (len(data["metadata"]) - 1),
            *[[k, v] for k, v in data["metadata"].items()],
            ["\nParamètres de Simulation"] + [""] * (len(data["configuration"]) - 1),
            *[[k, v] for k, v in data["configuration"].items()]
        ]
        
        results_data = [
            ["Résultats de la Simulation"] + [""] * (len(data["results"]) - 1),
            *[[k, v] for k, v in data["results"].items()]
        ]
        
        calculations_data = [
            ["Calculs Intermédiaires"] + [""] * (len(data["calculations"]) - 1),
            *[[k, v] for k, v in data["calculations"].items()]
        ]
        
        # Écrire les données dans les feuilles
        config_sheet.update('A1', config_data)
        results_sheet.update('A1', results_data)
        calculations_sheet.update('A1', calculations_data)
        
        # Ajouter des notes explicatives
        config_sheet.insert_note('A1', 'Configuration du projet - Ne pas modifier')
        results_sheet.insert_note('A1', 'Résultats de la simulation - Ne pas modifier')
        calculations_sheet.insert_note('A1', 'Calculs intermédiaires - Ne pas modifier')
        simulation_sheet.insert_note('A1', 'Feuille de simulation - Modifiez les valeurs ici')
        
        return spreadsheet.url
        
    except Exception as e:
        return f"Erreur lors de l'export vers Google Sheets: {str(e)}"

def export_to_excel(data, filename="simulation_rhum.xlsx"):
    """
    Exporte les données au format Excel (.xlsx) avec une structure claire et formatée
    """
    try:
        # Créer un nouveau workbook Excel
        wb = Workbook()
        
        # 1. Feuille de Configuration
        config_ws = wb.active
        config_ws.title = "Configuration"
        
        # Style pour les titres
        title_font = Font(bold=True, size=14)
        header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        
        # Ajouter les métadonnées
        config_ws.append(["Configuration Générale"])
        config_ws['A1'].font = title_font
        for key, value in data["metadata"].items():
            config_ws.append([key, value])
            config_ws.cell(row=config_ws.max_row, column=1).fill = header_fill
            config_ws.cell(row=config_ws.max_row, column=1).alignment = Alignment(horizontal="right")
        
        # Ajouter les paramètres de simulation
        config_ws.append([])
        config_ws.append(["Paramètres de Simulation"])
        config_ws['A' + str(config_ws.max_row)].font = title_font
        for key, value in data["configuration"].items():
            config_ws.append([key, value])
            config_ws.cell(row=config_ws.max_row, column=1).fill = header_fill
            config_ws.cell(row=config_ws.max_row, column=1).alignment = Alignment(horizontal="right")
        
        # 2. Feuille de Résultats
        results_ws = wb.create_sheet("Résultats")
        
        # Préparer les données de résultats
        results_data = [
            ["Production PV (serre)", data["results"]["production_pv"]],
            ["Production PV (au sol)", data["results"]["production_au_sol"]],
            ["Production totale", data["results"]["production_total"]],
            ["Autoconsommation", data["results"]["autoconsommation"]],
            ["Revente", data["results"]["revente"]],
            ["Revenu PV", data["results"]["revenu_pv"]],
            ["Revenu Rhum", data["results"]["revenu_rhum"]],
            ["Revenu total", data["results"]["revenu_total"]],
            ["Coût PV", data["results"]["cout_pv"]],
            ["Coût serre", data["results"]["cout_serre"]],
            ["Coût total", data["results"]["cout_total"]],
            ["Bénéfice net", data["results"]["benefice_net"]],
            ["ROI", data["results"]["roi"]],
            ["Temps retour", data["results"]["temps_retour"]]
        ]
        
        # Ajouter les données
        results_ws.append(["Résultats de la Simulation"])
        results_ws['A1'].font = title_font
        
        for row in results_data:
            results_ws.append(row)
            results_ws.cell(row=results_ws.max_row, column=1).fill = header_fill
            results_ws.cell(row=results_ws.max_row, column=1).alignment = Alignment(horizontal="right")
        
        # 3. Feuille de Production Mensuelle
        monthly_ws = wb.create_sheet("Production Mensuelle")
        
        # Préparer les données de production mensuelle
        monthly_data = list(data["monthly_production"].items())
        monthly_ws.append(["Mois", "Production (kWh)"])
        
        for month, production in monthly_data:
            monthly_ws.append([month, production])
        
        # Sauvegarder le fichier Excel
        wb.save(filename)
        
        return filename
        
    except Exception as e:
        return f"Erreur lors de l'export Excel: {str(e)}"

def export_to_json(data, filename="simulation_results.json"):
    """
    Exporte les résultats de la simulation au format JSON
    """
    try:
        with open(filename, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=4)
        return filename
    except Exception as e:
        return f"Erreur lors de l'export JSON: {str(e)}"
