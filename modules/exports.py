import os
import json
import pandas as pd
import zipfile
from openpyxl import Workbook
import tempfile
import shutil
from datetime import datetime
import streamlit as st

from modules.state_manager import rhuma, rhuma_label, rhuma_description, state_manager

def export_to_json(data, filename="simulation_results.json"):
    try:
        with open(filename, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=4)
        return filename
    except Exception as e:
        return f"Erreur : {e}"

def export_to_csv(data, filename_prefix="simulation_rhum"):
    try:
        df = pd.DataFrame(data)
        csv_filename = f"{filename_prefix}.csv"
        df.to_csv(csv_filename, index=False, encoding='utf-8')
        return csv_filename
    except Exception as e:
        return f"Erreur : {e}"

def export_to_excel(data, filename="simulation_results.xlsx"):
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Simulation Results"
        row = 1
        for key, value in data.items():
            ws.cell(row=row, column=1, value=key)
            ws.cell(row=row, column=2, value=str(value))
            row += 1
        wb.save(filename)
        return filename
    except Exception as e:
        return f"Erreur : {e}"

def export_all_formats(data, filename_prefix="simulation_rhum"):
    try:
        temp_dir = tempfile.mkdtemp()
        json_file = os.path.join(temp_dir, f"{filename_prefix}.json")
        with open(json_file, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=4)
        excel_file = os.path.join(temp_dir, f"{filename_prefix}.xlsx")
        export_to_excel(data, excel_file)
        csv_file = os.path.join(temp_dir, f"{filename_prefix}.csv")
        df = pd.DataFrame(data)
        df.to_csv(csv_file, index=False, encoding='utf-8')
        zip_filename = f"{filename_prefix}_export.zip"
        with zipfile.ZipFile(zip_filename, 'w') as zipf:
            zipf.write(json_file, os.path.basename(json_file))
            zipf.write(excel_file, os.path.basename(excel_file))
            zipf.write(csv_file, os.path.basename(csv_file))
        shutil.rmtree(temp_dir)
        return zip_filename
    except Exception as e:
        return f"Erreur : {e}"
    

# Boutons individuels pour chaque format
st.subheader("Exporter un format spécifique")

# Export CSV
if st.button("CSV", key="export_csv"):    
    # Utiliser directement les résultats du gestionnaire d'état
    data = state_manager.get_results()
    export_to_csv(data)

# Export Excel
if st.button("Excel", key="export_excel"):    
    # Utiliser directement les résultats du gestionnaire d'état
    data = state_manager.get_results()
    export_to_excel(data)

# Export JSON
if st.button("JSON", key="export_json"):    
    # Utiliser directement les résultats du gestionnaire d'état
    data = state_manager.get_results()
    export_to_json(data)

# Lien vers le dépôt GitHub
st.markdown("[GitHub Repository](https://github.com/JeanHuguesRobert/Rhuma)")

# Fonction d'export de configuration
def export_config(config_name):
    """
    Exporte les configurations personnalisées dans un fichier .env.config_name
    """
    try:
        # Vérifier le nom de la configuration
        if not config_name or not config_name.replace('_', '').isalnum():
            st.error("Le nom de la configuration doit contenir uniquement des lettres, chiffres et underscores")
            return

        # Créer le contenu du fichier
        config_content = "# Configuration personnalisée\n"
        
        # Ajouter les variables de configuration
        config = state_manager.get_configuration()
        for key, value in config.items():
            config_content += f"RHUMA_{key.upper()}={value}\n"
            
        # Ajouter les métadonnées
        metadata = state_manager.get_metadata()
        for key, value in metadata.items():
            if key not in ['timestamp'] and not key.startswith('RHUMA_GOOGLE_SHEETS_'):
                config_content += f"RHUMA_{key.upper()}={value}\n"

        # Écrire le fichier
        filename = f".env.{config_name}"
        with open(filename, 'w') as f:
            f.write(config_content)

        st.success(f"Configuration exportée avec succès dans {filename}")
        
        # Télécharger le fichier
        with open(filename, 'r') as f:
            st.download_button(
                label="Télécharger la configuration",
                data=f,
                file_name=filename,
                mime="text/plain"
            )

    except Exception as e:
        st.error(f"Erreur lors de l'export de la configuration: {str(e)}")

# Ajouter un séparateur
st.sidebar.markdown("---")

# Export de configuration
config_name = st.sidebar.text_input("Nom de la configuration", "")
if st.sidebar.button("Exporter la configuration"):
    export_config(config_name)

# Paramètres ID
st.sidebar.markdown("---")
st.sidebar.subheader("Paramètres Généraux")

# ID technique
id = st.sidebar.text_input(rhuma_label('id'), rhuma('id'), help=rhuma_description('id'))
if id != rhuma('id'):
    if not id.replace('_', '').isalnum():
        st.error("L'ID doit contenir uniquement des lettres, chiffres et underscores")
    else:
        state_manager.set("id", id)

# Label utilisateur
label = st.sidebar.text_input(rhuma_label('label'), rhuma('label'), help=rhuma_description('label'))
if label != rhuma('label'):
    state_manager.set("label", label)

def import_from_json(file_path):
    """
    Importe les données depuis un fichier JSON et gère les versions
    
    Args:
        file_path (str): Chemin vers le fichier JSON
        
    Returns:
        dict: Données importées et adaptées à la version courante
        str: Message d'erreur si l'import échoue
    """
    try:
        # Lire le fichier JSON
        with open(file_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
            
        # Vérifier la version
        file_version = data.get('metadata', {}).get('version', '1.0.0')
        current_version = "1.0.0"  # Version actuelle de l'application
        
        # Comparer les versions
        if file_version != current_version:
            st.warning(f"Attention : le fichier est en version {file_version} alors que l'application est en version {current_version}")
            
            # Adapter les données selon la version
            if file_version == "0.1.0":
                data = adapt_from_v0_1_0(data)
            elif file_version == "0.2.0":
                data = adapt_from_v0_2_0(data)
            # Ajouter d'autres cas de conversion selon les versions
            
        # Vérifier l'intégrité des données
        if not validate_data(data):
            return None, "Les données importées ne sont pas valides"
            
        return data, None
        
    except json.JSONDecodeError:
        return None, "Le fichier n'est pas un JSON valide"
    except Exception as e:
        return None, f"Erreur lors de l'import : {str(e)}"


def adapt_from_v0_1_0(data):
    """
    Adapte les données de la version 0.1.0 à la version actuelle
    """
    # Exemple de conversion
    if 'old_parameter' in data:
        data['new_parameter'] = data.pop('old_parameter')
    
    return data


def adapt_from_v0_2_0(data):
    """
    Adapte les données de la version 0.2.0 à la version actuelle
    """
        
    return data


def validate_data(data):
    """
    Vérifie l'intégrité des données importées
    """
    required_fields = [
        'metadata',
        'configuration',
        'calculs',
        'scenarios',
        'monthly_production'
    ]
    
    for field in required_fields:
        if field not in data:
            return False
    
    # Vérifier la structure des données
    if not isinstance(data['configuration'], dict):
        return False
    if not isinstance(data['calculs'], dict):
        return False
    
    return True

def export_to_json(data, filename="simulation_results.json"):
    """
    Exporte les résultats de la simulation au format JSON avec une structure optimisée pour les bases de données
    
    Args:
        data (dict): Données à exporter
        filename (str): Nom du fichier JSON
        
    Returns:
        dict: Données exportées
    """
    try:
        # Préparer les données pour l'export
        export_data = {
            "metadata": {
                "id": rhuma("id"),
                "label": rhuma("label"),
                "version": rhuma("version"),
                "timestamp": rhuma("timestamp"),
                "export_timestamp": datetime.now().isoformat()
            },
            "configuration": {
                "surface_canne": rhuma('surface_canne'),
                "rendement_canne": rhuma('rendement_canne'),
                "teneur_sucre": rhuma('teneur_sucre'),
                "efficacite_extraction": rhuma('efficacite_extraction'),
                "efficacite_distillation": rhuma('efficacite_distillation'),
                "pv_serre": rhuma('pv_serre'),
                "pv_sol": rhuma('pv_sol'),
                "tarif_s24": rhuma('tarif_s24'),
                "tva": rhuma('tva'),
                "cout_fixe": rhuma('cout_fixe'),
                "cout_tracking": rhuma('cout_tracking'),
                "cout_construction": rhuma('cout_construction'),
                "cout_maintenance": rhuma('cout_maintenance'),
                "cout_assurance": rhuma('cout_assurance'),
                "cout_production": rhuma('cout_production'),
                "tarif_heures_creuses": rhuma('tarif_heures_creuses'),
                "autoconsommation_fixe": rhuma('autoconsommation_fixe'),
                "autoconsommation_tracking": rhuma('autoconsommation_tracking'),
                "prix_rhum": rhuma('prix_rhum'),
                "pertes_pv": rhuma('pertes_pv'),
                "pertes_tracking": rhuma('pertes_tracking'),
                "precision_tracking": rhuma('precision_tracking'),
                "taux_interet": rhuma('taux_interet'),
                "duree_amortissement": rhuma('duree_amortissement')
            },
            "results": {
                "production": {
                    "pv_serre": float(data.get("production_pv", 0)),
                    "pv_sol": float(data.get("production_au_sol", 0)),
                    "total": float(data.get("production_pv", 0) + data.get("production_au_sol", 0)),
                    "autoconsommation": float(data.get("autoconsommation", 0)),
                    "revente": float(data.get("revente", 0))
                },
                "revenus": {
                    "pv": float(data.get("revenu_pv", 0)),
                    "rhum": float(data.get("revenu_rhum", 0)),
                    "total": float(data.get("revenu_pv", 0) + data.get("revenu_rhum", 0))
                },
                "costs": {
                    "initial": {
                        "pv": float(data.get("cout_pv", 0)),
                        "serre": float(data.get("cout_serre", 0)),
                        "total": float(data.get("cout_total", 0))
                    },
                    "annual": {
                        "maintenance": float(data.get("couts_annuels", {}).get("maintenance", 0)),
                        "insurance": float(data.get("couts_annuels", {}).get("assurance", 0)),
                        "production": float(data.get("couts_annuels", {}).get("production", 0))
                    }
                },
                "benefits": {
                    "net": float(data.get("benefice_net", 0)),
                    "roi": float(data.get("roi", 0)),
                    "return_time": float(data.get("temps_retour", 0))
                }
            },
            "monthly_production": {
                "data": [{
                    "month": month,
                    "production": float(production)
                } for month, production in data.get("monthly_production", {}).items()]
            },
            "scenarios": [{
                "id": idx + 1,
                "name": scenario.get("nom", ""),
                "description": scenario.get("description", ""),
                "value": float(scenario.get("valeur", 0)) if scenario.get("valeur") is not None else None
            } for idx, scenario in enumerate(data.get("scenarios", []))]
        }

        # Écrire le fichier JSON
        with open(filename, 'w', encoding='utf-8') as f:
            json.dump(export_data, f, indent=4, ensure_ascii=False)
        
        # Télécharger le fichier
        with open(filename, 'r') as f:
            st.download_button(
                label="Télécharger le fichier JSON",
                data=f.read(),
                file_name=filename,
                mime="application/json"
            )
        
        st.success(f"Données exportées avec succès dans {filename}")
        
        return export_data
        
    except Exception as e:
        st.error(f"Erreur lors de l'export JSON: {str(e)}")
        return None

def export_all_formats(data, filename_prefix="simulation_rhum"):
    """
    Exporte les données dans tous les formats disponibles (JSON, CSV, Excel, .env)
    et les regroupe dans un fichier ZIP unique
    
    Args:
        data (dict): Données à exporter
        filename_prefix (str): Préfixe pour les noms de fichiers
        
    Returns:
        dict: Données exportées
    """
    try:
        # Créer un dossier temporaire pour stocker les fichiers
        temp_dir = tempfile.mkdtemp()
        
        # 1. Export JSON
        json_filename = os.path.join(temp_dir, f"{filename_prefix}.json")
        with open(json_filename, 'w', encoding='utf-8') as f:
            json.dump(export_to_json(data), f, indent=4, ensure_ascii=False)
        
        # 2. Export Excel
        excel_filename = os.path.join(temp_dir, f"{filename_prefix}.xlsx")
        export_to_excel(data, excel_filename)
        
        # 3. Export CSV (dans un dossier séparé)
        csv_dir = os.path.join(temp_dir, "csv")
        os.makedirs(csv_dir, exist_ok=True)
        
        # Configuration
        config_df = pd.DataFrame([
            [rhuma_label("id"), rhuma("id")],
            [rhuma_label("label"), rhuma("label")],
            [rhuma_label("version"), rhuma("version")],
            ["Timestamp", rhuma("timestamp")]
        ])
        config_df.to_csv(os.path.join(csv_dir, f"{filename_prefix}_configuration.csv"), index=False, encoding='utf-8')
        
        # Paramètres
        params_df = pd.DataFrame([
            ["Surface canne (m²)", rhuma("surface_canne")],
            ["Rendement canne (t/ha)", rhuma("rendement_canne")],
            ["Teneur sucre (%)", rhuma("teneur_sucre")],
            ["Efficacité extraction (%)", rhuma("efficacite_extraction")],
            ["Efficacité distillation (%)", rhuma("efficacite_distillation")],
            ["Puissance PV (serre) (kWc)", rhuma("pv_serre")],
            ["Puissance PV (au sol) (kWc)", rhuma("pv_sol")],
            ["Tarif S24 (€/kWh)", rhuma("tarif_s24")],
            ["TVA (%)", rhuma("tva")],
            ["Coût système PV fixe (€/kWc)", rhuma("cout_fixe")],
            ["Coût système tracking (€/kWc)", rhuma("cout_tracking")],
            ["Coût construction serre (€/m²)", rhuma("cout_construction")],
            ["Coût maintenance (€/kWc)", rhuma("cout_maintenance")],
            ["Coût assurance (€/kWc)", rhuma("cout_assurance")],
            ["Coût production (€/kWc)", rhuma("cout_production")],
            ["Tarif heures creuses (€/kWh)", rhuma("tarif_heures_creuses")],
            ["Autoconsommation fixe (kWh)", rhuma("autoconsommation_fixe")],
            ["Autoconsommation tracking (kWh)", rhuma("autoconsommation_tracking")],
            ["Prix du rhum (€/L)", rhuma("prix_rhum")],
            ["Pertes PV (%)", rhuma("pertes_pv")],
            ["Pertes tracking (%)", rhuma("pertes_tracking")],
            ["Précision tracking (°)", rhuma("precision_tracking")],
            ["Taux d'intérêt (%)", rhuma("taux_interet")],
            ["Durée d'amortissement (ans)", rhuma("duree_amortissement")]
        ])
        params_df.to_csv(os.path.join(csv_dir, f"{filename_prefix}_parameters.csv"), index=False, encoding='utf-8')
        
        # Résultats
        results_df = pd.DataFrame([
            ["Production PV (serre)", data.get("production_pv", 0)],
            ["Production PV (au sol)", data.get("production_au_sol", 0)],
            ["Production totale", data.get("production_pv", 0) + data.get("production_au_sol", 0)],
            ["Autoconsommation", data.get("autoconsommation", 0)],
            ["Revente", data.get("revente", 0)],
            ["Revenu PV", data.get("revenu_pv", 0)],
            ["Revenu Rhum", data.get("revenu_rhum", 0)],
            ["Revenu total", data.get("revenu_pv", 0) + data.get("revenu_rhum", 0)],
            ["Coût PV", data.get("cout_pv", 0)],
            ["Coût serre", data.get("cout_serre", 0)],
            ["Coût total", data.get("cout_total", 0)],
            ["Bénéfice net", data.get("benefice_net", 0)],
            ["ROI", data.get("roi", 0)],
            ["Temps retour", data.get("temps_retour", 0)]
        ])
        results_df.to_csv(os.path.join(csv_dir, f"{filename_prefix}_results.csv"), index=False, encoding='utf-8')
        
        # Production mensuelle
        monthly_df = pd.DataFrame(list(data.get("monthly_production", {}).items()),
                                columns=["Mois", "Production (kWh)"])
        monthly_df.to_csv(os.path.join(csv_dir, f"{filename_prefix}_monthly_production.csv"), index=False, encoding='utf-8')
        
        # 4. Export .env (sans les secrets)
        env_filename = os.path.join(temp_dir, f"{filename_prefix}.env")
        with open(env_filename, 'w', encoding='utf-8') as f:
            f.write(f"RHUMA_ID={rhuma('id')}\n")
            f.write(f"RHUMA_LABEL={rhuma('label')}\n")
            f.write(f"RHUMA_VERSION={rhuma('version')}\n")
            # Récupérer la configuration complète
            config = state_manager.get_configuration()
            for key, value in config.items():
                if isinstance(value, (int, float)):
                    f.write(f"RHUMA_{key.upper()}={value}\n")
        
        # Créer le fichier ZIP
        zip_filename = f"{filename_prefix}_export.zip"
        with zipfile.ZipFile(zip_filename, 'w') as zipf:
            # Ajouter tous les fichiers
            zipf.write(json_filename, os.path.basename(json_filename))
            zipf.write(excel_filename, os.path.basename(excel_filename))
            zipf.write(env_filename, os.path.basename(env_filename))
            
            # Ajouter les fichiers CSV
            for csv_file in os.listdir(csv_dir):
                zipf.write(os.path.join(csv_dir, csv_file), f"csv/{csv_file}")
        
        # Nettoyer les fichiers temporaires
        shutil.rmtree(temp_dir)
        
        # Télécharger le fichier ZIP
        with open(zip_filename, 'rb') as f:
            st.download_button(
                label="Télécharger l'export complet",
                data=f,
                file_name=zip_filename,
                mime="application/zip"
            )
        
        st.success("Données exportées avec succès dans tous les formats")
        
        return data
        
    except Exception as e:
        st.error(f"Erreur lors de l'export complet: {str(e)}")
        return None

# Export des résultats
if st.button("💾 Exporter en CSV"):
    # Utiliser directement les résultats du gestionnaire d'état
    results = state_manager.get_results()
    monthly_production = results.get("monthly_production", {})
    monthly_pv_production = results.get("monthly_pv_production", {})
    monthly_au_sol_production = results.get("monthly_au_sol_production", {})
    production_pv = results.get("production_pv", 0)
    production_au_sol = results.get("production_au_sol", 0)
    chiffre_affaires_collectif = results.get("chiffre_affaires_collectif", 0)
    
    data = {
        "Mois": list(monthly_production.keys()) + ["Total annuel"],
        "Production serre (MWh)": [x/1000 for x in list(monthly_pv_production.values())] + [production_pv/1000],
        "Production collective (MWh)": [x/1000 for x in list(monthly_au_sol_production.values())] + [production_au_sol/1000],
        "CA collectif (€)": [chiffre_affaires_collectif] * 12 + [chiffre_affaires_collectif]
    }
    df_export = pd.DataFrame(data)
    st.download_button("⬇️ Télécharger", df_export.to_csv(index=False), "production_rhum_solaire.csv", "text/csv")

def export_to_csv(data, filename_prefix="simulation_rhum"):
    """
    Exporte les données au format CSV avec une structure claire
    
    Args:
        data (dict): Données à exporter
        filename_prefix (str): Préfixe du nom du fichier
        
    Returns:
        dict: Données exportées
    """
    try:
        # Préparer les données pour l'export
        export_data = {
            "metadata": {
                "id": rhuma("id"),
                "label": rhuma("label"),
                "version": rhuma("version"),
                "timestamp": rhuma("timestamp")
            },
            "configuration": state_manager.get_configuration(),
            "calculs": {
                "production": {
                    "pv_serre": data.get("production_pv", 0),
                    "pv_sol": data.get("production_au_sol", 0),
                    "total": data.get("production_pv", 0) + data.get("production_au_sol", 0),
                    "autoconsommation": data.get("autoconsommation", 0),
                    "revente": data.get("revente", 0)
                },
                "revenus": {
                    "pv": data.get("revenu_pv", 0),
                    "rhum": data.get("revenu_rhum", 0),
                    "total": data.get("revenu_pv", 0) + data.get("revenu_rhum", 0)
                },
                "couts": {
                    "initiaux": {
                        "pv": data.get("cout_pv", 0),
                        "serre": data.get("cout_serre", 0),
                        "total": data.get("cout_total", 0)
                    },
                    "annuels": {
                        "maintenance": data.get("couts_annuels", {}).get("maintenance", 0),
                        "assurance": data.get("couts_annuels", {}).get("assurance", 0),
                        "production": data.get("couts_annuels", {}).get("production", 0)
                    }
                },
                "benefices": {
                    "net": data.get("benefice_net", 0),
                    "roi": data.get("roi", 0),
                    "temps_retour": data.get("temps_retour", 0)
                }
            },
            "scenarios": data.get("scenarios", []),
            "monthly_production": data.get("monthly_production", {})
        }
        
        # Créer les fichiers CSV pour chaque feuille
        filenames = {}
        
        # 1. Configuration
        config_df = pd.DataFrame([
            ["ID du Projet", rhuma("id")],
            ["Label du Projet", rhuma("label")],
            ["Version", rhuma("version")],
            ["Timestamp", rhuma("timestamp")]
        ])
        config_df.to_csv(f"{filename_prefix}_configuration.csv", index=False, encoding='utf-8')
        filenames["configuration"] = f"{filename_prefix}_configuration.csv"
        
        # 2. Paramètres de Simulation
        params_df = pd.DataFrame([
            ["Surface canne (m²)", rhuma("surface_canne")],
            ["Rendement canne (t/ha)", rhuma("rendement_canne")],
            ["Teneur sucre (%)", rhuma("teneur_sucre")],
            ["Efficacité extraction (%)", rhuma("efficacite_extraction")],
            ["Efficacité distillation (%)", rhuma("efficacite_distillation")],
            ["Puissance PV (serre) (kWc)", rhuma("pv_serre")],
            ["Puissance PV (au sol) (kWc)", rhuma("pv_sol")],
            ["Tarif S24 (€/kWh)", rhuma("tarif_s24")],
            ["TVA (%)", rhuma("tva")],
            ["Coût système PV fixe (€/kWc)", rhuma("cout_fixe")],
            ["Coût système tracking (€/kWc)", rhuma("cout_tracking")],
            ["Coût construction serre (€/m²)", rhuma("cout_construction")],
            ["Coût maintenance (€/kWc)", rhuma("cout_maintenance")],
            ["Coût assurance (€/kWc)", rhuma("cout_assurance")],
            ["Coût production (€/kWc)", rhuma("cout_production")],
            ["Tarif heures creuses (€/kWh)", rhuma("tarif_heures_creuses")],
            ["Autoconsommation fixe (kWh)", rhuma("autoconsommation_fixe")],
            ["Autoconsommation tracking (kWh)", rhuma("autoconsommation_tracking")],
            ["Prix du rhum (€/L)", rhuma("prix_rhum")],
            ["Pertes PV (%)", rhuma("pertes_pv")],
            ["Pertes tracking (%)", rhuma("pertes_tracking")],
            ["Précision tracking (°)", rhuma("precision_tracking")],
            ["Taux d'intérêt (%)", rhuma("taux_interet")],
            ["Durée d'amortissement (ans)", rhuma("duree_amortissement")]
        ])
        params_df.to_csv(f"{filename_prefix}_parameters.csv", index=False, encoding='utf-8')
        filenames["parameters"] = f"{filename_prefix}_parameters.csv"
        
        # 3. Résultats
        results_df = pd.DataFrame([
            ["Production PV (serre)", data.get("production_pv", 0)],
            ["Production PV (au sol)", data.get("production_au_sol", 0)],
            ["Production totale", data.get("production_pv", 0) + data.get("production_au_sol", 0)],
            ["Autoconsommation", data.get("autoconsommation", 0)],
            ["Revente", data.get("revente", 0)],
            ["Revenu PV", data.get("revenu_pv", 0)],
            ["Revenu Rhum", data.get("revenu_rhum", 0)],
            ["Revenu total", data.get("revenu_pv", 0) + data.get("revenu_rhum", 0)],
            ["Coût PV", data.get("cout_pv", 0)],
            ["Coût serre", data.get("cout_serre", 0)],
            ["Coût total", data.get("cout_total", 0)],
            ["Bénéfice net", data.get("benefice_net", 0)],
            ["ROI", data.get("roi", 0)],
            ["Temps retour", data.get("temps_retour", 0)]
        ])
        results_df.to_csv(f"{filename_prefix}_results.csv", index=False, encoding='utf-8')
        filenames["results"] = f"{filename_prefix}_results.csv"
        
        # 4. Production mensuelle
        monthly_df = pd.DataFrame(list(data.get("monthly_production", {}).items()),
                                columns=["Mois", "Production (kWh)"])
        monthly_df.to_csv(f"{filename_prefix}_monthly_production.csv", index=False, encoding='utf-8')
        filenames["monthly_production"] = f"{filename_prefix}_monthly_production.csv"
        
        # Créer un fichier ZIP avec tous les fichiers CSV
        zip_filename = f"{filename_prefix}_export.zip"
        with zipfile.ZipFile(zip_filename, 'w') as zipf:
            # Ajouter tous les fichiers
            zipf.write(f"{filename_prefix}_configuration.csv", os.path.basename(f"{filename_prefix}_configuration.csv"))
            zipf.write(f"{filename_prefix}_parameters.csv", os.path.basename(f"{filename_prefix}_parameters.csv"))
            zipf.write(f"{filename_prefix}_results.csv", os.path.basename(f"{filename_prefix}_results.csv"))
            zipf.write(f"{filename_prefix}_monthly_production.csv", os.path.basename(f"{filename_prefix}_monthly_production.csv"))
            os.remove(f"{filename_prefix}_configuration.csv")
            os.remove(f"{filename_prefix}_parameters.csv")
            os.remove(f"{filename_prefix}_results.csv")
            os.remove(f"{filename_prefix}_monthly_production.csv")
        
        # Télécharger le fichier ZIP
        with open(zip_filename, 'rb') as f:
            st.download_button(
                label="Télécharger l'export CSV",
                data=f,
                file_name=zip_filename,
                mime="application/zip"
            )
        
        st.success("Données exportées avec succès au format CSV")
        
        return export_data
        
    except Exception as e:
        st.error(f"Erreur lors de l'export CSV: {str(e)}")
        return None

def export_to_json(data, filename="simulation_results.json"):
    """
    Exporte les résultats de la simulation au format JSON avec une structure optimisée pour les bases de données
    
    Args:
        data (dict): Données à exporter
        filename (str): Nom du fichier JSON
        
    Returns:
        dict: Données exportées
    """
    try:
        # Préparer les données pour l'export
        export_data = {
            "metadata": {
                "id": rhuma("id"),
                "label": rhuma("label"),
                "version": rhuma("version"),
                "timestamp": rhuma("timestamp"),
                "export_timestamp": datetime.now().isoformat()
            },
            "configuration": {
                "surface_canne": rhuma('surface_canne'),
                "rendement_canne": rhuma('rendement_canne'),
                "teneur_sucre": rhuma('teneur_sucre'),
                "efficacite_extraction": rhuma('efficacite_extraction'),
                "efficacite_distillation": rhuma('efficacite_distillation'),
                "pv_serre": rhuma('pv_serre'),
                "pv_sol": rhuma('pv_sol'),
                "tarif_s24": rhuma('tarif_s24'),
                "tva": rhuma('tva'),
                "cout_fixe": rhuma('cout_fixe'),
                "cout_tracking": rhuma('cout_tracking'),
                "cout_construction": rhuma('cout_construction'),
                "cout_maintenance": rhuma('cout_maintenance'),
                "cout_assurance": rhuma('cout_assurance'),
                "cout_production": rhuma('cout_production'),
                "tarif_heures_creuses": rhuma('tarif_heures_creuses'),
                "autoconsommation_fixe": rhuma('autoconsommation_fixe'),
                "autoconsommation_tracking": rhuma('autoconsommation_tracking'),
                "prix_rhum": rhuma('prix_rhum'),
                "pertes_pv": rhuma('pertes_pv'),
                "pertes_tracking": rhuma('pertes_tracking'),
                "precision_tracking": rhuma('precision_tracking'),
                "taux_interet": rhuma('taux_interet'),
                "duree_amortissement": rhuma('duree_amortissement')
            },
            "results": {
                "production": {
                    "pv_serre": float(data.get("production_pv", 0)),
                    "pv_sol": float(data.get("production_au_sol", 0)),
                    "total": float(data.get("production_pv", 0) + data.get("production_au_sol", 0)),
                    "autoconsommation": float(data.get("autoconsommation", 0)),
                    "revente": float(data.get("revente", 0))
                },
                "revenus": {
                    "pv": float(data.get("revenu_pv", 0)),
                    "rhum": float(data.get("revenu_rhum", 0)),
                    "total": float(data.get("revenu_pv", 0) + data.get("revenu_rhum", 0))
                },
                "costs": {
                    "initial": {
                        "pv": float(data.get("cout_pv", 0)),
                        "serre": float(data.get("cout_serre", 0)),
                        "total": float(data.get("cout_total", 0))
                    },
                    "annual": {
                        "maintenance": float(data.get("couts_annuels", {}).get("maintenance", 0)),
                        "insurance": float(data.get("couts_annuels", {}).get("assurance", 0)),
                        "production": float(data.get("couts_annuels", {}).get("production", 0))
                    }
                },
                "benefits": {
                    "net": float(data.get("benefice_net", 0)),
                    "roi": float(data.get("roi", 0)),
                    "return_time": float(data.get("temps_retour", 0))
                }
            },
            "monthly_production": {
                "data": [{
                    "month": month,
                    "production": float(production)
                } for month, production in data.get("monthly_production", {}).items()]
            },
            "scenarios": [{
                "id": idx + 1,
                "name": scenario.get("nom", ""),
                "description": scenario.get("description", ""),
                "value": float(scenario.get("valeur", 0)) if scenario.get("valeur") is not None else None
            } for idx, scenario in enumerate(data.get("scenarios", []))]
        }

        # Écrire le fichier JSON
        with open(filename, 'w', encoding='utf-8') as f:
            json.dump(export_data, f, indent=4, ensure_ascii=False)
        
        # Télécharger le fichier
        with open(filename, 'r') as f:
            st.download_button(
                label="Télécharger le fichier JSON",
                data=f.read(),
                file_name=filename,
                mime="application/json"
            )
        
        st.success(f"Données exportées avec succès dans {filename}")
        
        return export_data
        
    except Exception as e:
        st.error(f"Erreur lors de l'export JSON: {str(e)}")
        return None

def export_all_formats(data, filename_prefix="simulation_rhum"):
    """
    Exporte les données dans tous les formats disponibles (JSON, CSV, Excel, .env)
    et les regroupe dans un fichier ZIP unique
    
    Args:
        data (dict): Données à exporter
        filename_prefix (str): Préfixe pour les noms de fichiers
        
    Returns:
        dict: Données exportées
    """
    try:
        # Créer un dossier temporaire pour stocker les fichiers
        temp_dir = tempfile.mkdtemp()
        
        # 1. Export JSON
        json_filename = os.path.join(temp_dir, f"{filename_prefix}.json")
        with open(json_filename, 'w', encoding='utf-8') as f:
            json.dump(export_to_json(data), f, indent=4, ensure_ascii=False)
        
        # 2. Export Excel
        excel_filename = os.path.join(temp_dir, f"{filename_prefix}.xlsx")
        export_to_excel(data, excel_filename)
        
        # 3. Export CSV (dans un dossier séparé)
        csv_dir = os.path.join(temp_dir, "csv")
        os.makedirs(csv_dir, exist_ok=True)
        
        # Configuration
        config_df = pd.DataFrame([
            [rhuma_label("id"), rhuma("id")],
            [rhuma_label("label"), rhuma("label")],
            [rhuma_label("version"), rhuma("version")],
            ["Timestamp", rhuma("timestamp")]
        ])
        config_df.to_csv(os.path.join(csv_dir, f"{filename_prefix}_configuration.csv"), index=False, encoding='utf-8')
        
        # Paramètres
        params_df = pd.DataFrame([
            ["Surface canne (m²)", rhuma("surface_canne")],
            ["Rendement canne (t/ha)", rhuma("rendement_canne")],
            ["Teneur sucre (%)", rhuma("teneur_sucre")],
            ["Efficacité extraction (%)", rhuma("efficacite_extraction")],
            ["Efficacité distillation (%)", rhuma("efficacite_distillation")],
            ["Puissance PV (serre) (kWc)", rhuma("pv_serre")],
            ["Puissance PV (au sol) (kWc)", rhuma("pv_sol")],
            ["Tarif S24 (€/kWh)", rhuma("tarif_s24")],
            ["TVA (%)", rhuma("tva")],
            ["Coût système PV fixe (€/kWc)", rhuma("cout_fixe")],
            ["Coût système tracking (€/kWc)", rhuma("cout_tracking")],
            ["Coût construction serre (€/m²)", rhuma("cout_construction")],
            ["Coût maintenance (€/kWc)", rhuma("cout_maintenance")],
            ["Coût assurance (€/kWc)", rhuma("cout_assurance")],
            ["Coût production (€/kWc)", rhuma("cout_production")],
            ["Tarif heures creuses (€/kWh)", rhuma("tarif_heures_creuses")],
            ["Autoconsommation fixe (kWh)", rhuma("autoconsommation_fixe")],
            ["Autoconsommation tracking (kWh)", rhuma("autoconsommation_tracking")],
            ["Prix du rhum (€/L)", rhuma("prix_rhum")],
            ["Pertes PV (%)", rhuma("pertes_pv")],
            ["Pertes tracking (%)", rhuma("pertes_tracking")],
            ["Précision tracking (°)", rhuma("precision_tracking")],
            ["Taux d'intérêt (%)", rhuma("taux_interet")],
            ["Durée d'amortissement (ans)", rhuma("duree_amortissement")]
        ])
        params_df.to_csv(os.path.join(csv_dir, f"{filename_prefix}_parameters.csv"), index=False, encoding='utf-8')
        
        # Résultats
        results_df = pd.DataFrame([
            ["Production PV (serre)", data.get("production_pv", 0)],
            ["Production PV (au sol)", data.get("production_au_sol", 0)],
            ["Production totale", data.get("production_pv", 0) + data.get("production_au_sol", 0)],
            ["Autoconsommation", data.get("autoconsommation", 0)],
            ["Revente", data.get("revente", 0)],
            ["Revenu PV", data.get("revenu_pv", 0)],
            ["Revenu Rhum", data.get("revenu_rhum", 0)],
            ["Revenu total", data.get("revenu_pv", 0) + data.get("revenu_rhum", 0)],
            ["Coût PV", data.get("cout_pv", 0)],
            ["Coût serre", data.get("cout_serre", 0)],
            ["Coût total", data.get("cout_total", 0)],
            ["Bénéfice net", data.get("benefice_net", 0)],
            ["ROI", data.get("roi", 0)],
            ["Temps retour", data.get("temps_retour", 0)]
        ])
        results_df.to_csv(os.path.join(csv_dir, f"{filename_prefix}_results.csv"), index=False, encoding='utf-8')
        
        # Production mensuelle
        monthly_df = pd.DataFrame(list(data.get("monthly_production", {}).items()),
                                columns=["Mois", "Production (kWh)"])
        monthly_df.to_csv(os.path.join(csv_dir, f"{filename_prefix}_monthly_production.csv"), index=False, encoding='utf-8')
        
        # 4. Export .env (sans les secrets)
        env_filename = os.path.join(temp_dir, f"{filename_prefix}.env")
        with open(env_filename, 'w', encoding='utf-8') as f:
            f.write(f"RHUMA_ID={rhuma('id')}\n")
            f.write(f"RHUMA_LABEL={rhuma('label')}\n")
            f.write(f"RHUMA_VERSION={rhuma('version')}\n")
            # Récupérer la configuration complète
            config = state_manager.get_configuration()
            for key, value in config.items():
                if isinstance(value, (int, float)):
                    f.write(f"RHUMA_{key.upper()}={value}\n")
        
        # Créer le fichier ZIP
        zip_filename = f"{filename_prefix}_export.zip"
        with zipfile.ZipFile(zip_filename, 'w') as zipf:
            # Ajouter tous les fichiers
            zipf.write(json_filename, os.path.basename(json_filename))
            zipf.write(excel_filename, os.path.basename(excel_filename))
            zipf.write(env_filename, os.path.basename(env_filename))
            
            # Ajouter les fichiers CSV
            for csv_file in os.listdir(csv_dir):
                zipf.write(os.path.join(csv_dir, csv_file), f"csv/{csv_file}")
        
        # Nettoyer les fichiers temporaires
        shutil.rmtree(temp_dir)
        
        # Télécharger le fichier ZIP
        with open(zip_filename, 'rb') as f:
            st.download_button(
                label="Télécharger l'export complet",
                data=f,
                file_name=zip_filename,
                mime="application/zip"
            )
        
        st.success("Données exportées avec succès dans tous les formats")
        
        return data
        
    except Exception as e:
        st.error(f"Erreur lors de l'export complet: {str(e)}")
        return None
